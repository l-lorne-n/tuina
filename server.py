from __future__ import annotations

import argparse
import base64
import csv
import datetime as dt
import hashlib
import hmac
import io
import json
import mimetypes
import os
import re
import socketserver
import sqlite3
import sys
import threading
import time
import urllib.error
import urllib.parse
import urllib.request
import uuid
import wave
import zipfile
from http import HTTPStatus
from http.server import SimpleHTTPRequestHandler
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET


APP_DIR = Path(__file__).resolve().parent
STATIC_DIR = APP_DIR / "static"
DATA_DIR = APP_DIR / "data"
DB_PATH = DATA_DIR / "tuina_records.sqlite3"
EXCEL_DIR = APP_DIR / "doc_patient"
SIGNATURE_PATIENT_DIR = APP_DIR / "doc_patient" / "signature_patient"
SIGNATURE_MANIFEST_PATH = SIGNATURE_PATIENT_DIR / "manifest.json"
SIGNATURE_BINDINGS_PATH = SIGNATURE_PATIENT_DIR / "bindings.json"
ELECTRONIC_SIGNATURE_DIR = SIGNATURE_PATIENT_DIR / "electronic"
TEST_SIGNATURE_DIR = SIGNATURE_PATIENT_DIR / "electronic_test"
BACKUP_STATE_PATH = DATA_DIR / "backup_state.json"
BACKUP_TEMP_DIR = DATA_DIR / "backup_tmp"

TENCENT_ENDPOINT = "https://asr.tencentcloudapi.com"
TENCENT_HOST = "asr.tencentcloudapi.com"
TENCENT_SERVICE = "asr"
TENCENT_ACTION = "SentenceRecognition"
TENCENT_VERSION = "2019-06-14"

MAX_JSON_BYTES = 8 * 1024 * 1024
MAX_AUDIO_BYTES = 5 * 1024 * 1024
MAX_AUDIO_SECONDS = 60.5
MAX_SIGNATURE_IMAGE_BYTES = 5 * 1024 * 1024
DEFAULT_SESSION_PRICE = 90
THERAPISTS = ("王师傅", "杨师傅")
PATIENT_ADDRESS_OPTIONS = (
    "浑南区",
    "和平区",
    "大东区",
    "铁西区",
    "中海",
    "鹿特丹",
    "沈河区",
    "长白岛",
    "龙湖",
    "沈北新区",
    "苏家屯",
    "于洪区",
    "金沙湾",
    "外阜",
    "新加坡城",
    "皇姑区",
)
CORRECTION_REASONS = ("手误", "患者反悔", "其他")
BACKUP_RETRY_SECONDS = 60
BACKUP_RETENTION_DAYS = 30
BACKUP_MAGIC = b"TUINABAK1"
BACKUP_STATUS_LOCK = threading.Lock()
BACKUP_RUN_LOCK = threading.Lock()
BACKUP_WORKER_STARTED = False
BACKUP_STATUS: dict[str, Any] = {
    "state": "idle",
    "message": "云备份待检查",
    "configured": False,
    "updatedAt": "",
}

NUMBER_TOKEN = r"[0-9]+(?:\.[0-9]+)?|[零〇一二两三四五六七八九十百千万幺点]+"
NUM = f"(?:{NUMBER_TOKEN})"
LABEL_SEP = r"[\s,，.。;；:：、]*"
DATE_FRAGMENT = (
    fr"(?:{NUM}年{NUM}月{NUM}(?:日|号)?"
    fr"|[0-9]{{5,6}}月[0-9]{{1,2}}(?:日|号)?"
    fr"|{NUM}月{NUM}(?:日|号)?)"
)
CN_DIGITS = {
    "零": 0,
    "〇": 0,
    "一": 1,
    "幺": 1,
    "二": 2,
    "两": 2,
    "三": 3,
    "四": 4,
    "五": 5,
    "六": 6,
    "七": 7,
    "八": 8,
    "九": 9,
}
CN_UNITS = {"十": 10, "百": 100, "千": 1000, "万": 10000}


def load_dotenv() -> None:
    for env_path in (APP_DIR / ".env",):
        if not env_path.exists():
            continue
        for raw_line in env_path.read_text(encoding="utf-8").splitlines():
            line = raw_line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            if key and key not in os.environ:
                os.environ[key] = value


def load_tencent_key_file() -> None:
    for key_path in (
        APP_DIR / "tecent_api_key.txt",
        APP_DIR / "tencent_api_key.txt",
    ):
        if not key_path.exists():
            continue

        parsed: dict[str, str] = {}
        for raw_line in key_path.read_text(encoding="utf-8").splitlines():
            line = raw_line.strip()
            if not line or line.startswith("#"):
                continue
            if ":" in line:
                key, value = line.split(":", 1)
            elif "=" in line:
                key, value = line.split("=", 1)
            else:
                continue
            parsed[key.strip().lower()] = value.strip().strip('"').strip("'")

        if parsed.get("secretid") and "TENCENT_SECRET_ID" not in os.environ:
            os.environ["TENCENT_SECRET_ID"] = parsed["secretid"]
        if parsed.get("secretkey") and "TENCENT_SECRET_KEY" not in os.environ:
            os.environ["TENCENT_SECRET_KEY"] = parsed["secretkey"]
        return


def load_backup_key_file() -> None:
    key_path = APP_DIR / "jianguoyun_key.txt"
    if not key_path.exists():
        return

    parsed: dict[str, str] = {}
    for raw_line in key_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        parsed[key.strip()] = value.strip().strip('"').strip("'")

    for key, value in parsed.items():
        if key.startswith("BACKUP_") and key not in os.environ:
            os.environ[key] = value


def json_response(handler: SimpleHTTPRequestHandler, status: int, payload: dict[str, Any]) -> None:
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json; charset=utf-8")
    handler.send_header("Content-Length", str(len(body)))
    handler.send_header("Cache-Control", "no-store")
    handler.end_headers()
    handler.wfile.write(body)


def read_json_body(handler: SimpleHTTPRequestHandler, max_bytes: int = MAX_JSON_BYTES) -> dict[str, Any]:
    content_length = int(handler.headers.get("Content-Length", "0"))
    if content_length <= 0 or content_length > max_bytes:
        raise ValueError("请求体为空或过大。")
    body = handler.rfile.read(content_length)
    parsed = json.loads(body.decode("utf-8"))
    if not isinstance(parsed, dict):
        raise ValueError("请求体必须是 JSON 对象。")
    return parsed


def now_text() -> str:
    return dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def connect_db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_database() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with connect_db() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS patients (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                import_order INTEGER NOT NULL,
                source_area TEXT,
                source_seq TEXT,
                original_name TEXT NOT NULL,
                name TEXT NOT NULL,
                gender TEXT,
                age TEXT,
                weight TEXT,
                height TEXT,
                address TEXT,
                record_no INTEGER,
                phone TEXT,
                remaining_sessions INTEGER,
                raw_transcript TEXT,
                notes TEXT,
                status TEXT NOT NULL DEFAULT 'pending',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS recharges (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                patient_id INTEGER NOT NULL,
                position INTEGER NOT NULL,
                recharge_date TEXT,
                amount INTEGER,
                sessions INTEGER,
                raw_text TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY (patient_id) REFERENCES patients(id) ON DELETE CASCADE
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS session_adjustments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                patient_id INTEGER NOT NULL,
                operation TEXT NOT NULL,
                sessions INTEGER NOT NULL,
                amount REAL,
                therapist TEXT,
                before_sessions INTEGER,
                after_sessions INTEGER NOT NULL,
                signature_status TEXT,
                signature_url TEXT,
                signature_saved_at TEXT,
                signature_signer TEXT,
                signature_note TEXT,
                signature_batch_id INTEGER,
                voided_at TEXT,
                voided_by_adjustment_id INTEGER,
                correction_of_adjustment_id INTEGER,
                correction_reason TEXT,
                correction_note TEXT,
                occurred_at TEXT NOT NULL,
                note TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY (patient_id) REFERENCES patients(id) ON DELETE CASCADE
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS signature_batches (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                signature_url TEXT NOT NULL,
                signer TEXT,
                note TEXT,
                adjustment_count INTEGER NOT NULL,
                signed_at TEXT NOT NULL,
                created_at TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS settlements (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                start_date TEXT NOT NULL,
                end_date TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'active',
                snapshot_json TEXT NOT NULL,
                created_at TEXT NOT NULL,
                revoked_at TEXT,
                revoked_reason TEXT
            )
            """
        )
        ensure_patient_columns(conn)
        ensure_session_adjustment_columns(conn)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_patients_order ON patients(import_order)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_patients_name ON patients(name)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_patients_status ON patients(status)")
        conn.execute(
            """
            CREATE UNIQUE INDEX IF NOT EXISTS idx_patients_address_record_no_unique
            ON patients(COALESCE(address, ''), record_no)
            WHERE record_no IS NOT NULL
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS idx_recharges_patient ON recharges(patient_id)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_recharges_date ON recharges(recharge_date)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_session_adjustments_patient ON session_adjustments(patient_id)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_session_adjustments_occurred ON session_adjustments(occurred_at)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_session_adjustments_therapist ON session_adjustments(therapist)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_session_adjustments_signature_status ON session_adjustments(signature_status)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_session_adjustments_voided ON session_adjustments(voided_at)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_session_adjustments_correction ON session_adjustments(correction_of_adjustment_id)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_session_adjustments_signature_batch ON session_adjustments(signature_batch_id)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_signature_batches_signed ON signature_batches(signed_at)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_settlements_status_end ON settlements(status, end_date)")
        seed_patients_from_excel(conn)
        conn.commit()


def seed_patients_from_excel(conn: sqlite3.Connection) -> None:
    row = conn.execute("SELECT COUNT(*) AS count FROM patients").fetchone()
    if row and row["count"]:
        return

    excel_files = [
        path
        for path in sorted(EXCEL_DIR.glob("*.xlsx"))
        if not path.name.startswith("~$")
    ]
    if not excel_files:
        return

    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("缺少 openpyxl，无法导入人名 Excel。请先运行：python -m pip install openpyxl") from exc

    workbook = openpyxl.load_workbook(excel_files[0])
    sheet = workbook[workbook.sheetnames[0]]
    created_at = now_text()
    import_order = 1
    for row_cells in sheet.iter_rows(min_row=4, values_only=False):
        area = clean_cell(row_cells[0].value)
        seq = clean_cell(row_cells[1].value)
        name = clean_cell(row_cells[2].value)
        note = clean_cell(row_cells[3].value)
        name_struck = bool(getattr(row_cells[2].font, "strike", False))
        note_struck = bool(getattr(row_cells[3].font, "strike", False))

        if not name:
            continue
        if name_struck or note_struck or "划掉" in note or "删除" in note:
            continue

        conn.execute(
            """
            INSERT INTO patients (
                import_order, source_area, source_seq, original_name, name,
                status, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, 'pending', ?, ?)
            """,
            (import_order, area, seq, name, name, created_at, created_at),
        )
        import_order += 1


def clean_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def ensure_patient_columns(conn: sqlite3.Connection) -> None:
    columns = {row["name"] for row in conn.execute("PRAGMA table_info(patients)").fetchall()}
    if "record_no" not in columns:
        conn.execute("ALTER TABLE patients ADD COLUMN record_no INTEGER")


def ensure_session_adjustment_columns(conn: sqlite3.Connection) -> None:
    columns = {
        row["name"]
        for row in conn.execute("PRAGMA table_info(session_adjustments)").fetchall()
    }
    if "amount" not in columns:
        conn.execute("ALTER TABLE session_adjustments ADD COLUMN amount REAL")
    if "therapist" not in columns:
        conn.execute("ALTER TABLE session_adjustments ADD COLUMN therapist TEXT")
    if "signature_status" not in columns:
        conn.execute("ALTER TABLE session_adjustments ADD COLUMN signature_status TEXT")
    if "signature_url" not in columns:
        conn.execute("ALTER TABLE session_adjustments ADD COLUMN signature_url TEXT")
    if "signature_saved_at" not in columns:
        conn.execute("ALTER TABLE session_adjustments ADD COLUMN signature_saved_at TEXT")
    if "signature_signer" not in columns:
        conn.execute("ALTER TABLE session_adjustments ADD COLUMN signature_signer TEXT")
    if "signature_note" not in columns:
        conn.execute("ALTER TABLE session_adjustments ADD COLUMN signature_note TEXT")
    if "signature_batch_id" not in columns:
        conn.execute("ALTER TABLE session_adjustments ADD COLUMN signature_batch_id INTEGER")
    if "voided_at" not in columns:
        conn.execute("ALTER TABLE session_adjustments ADD COLUMN voided_at TEXT")
    if "voided_by_adjustment_id" not in columns:
        conn.execute("ALTER TABLE session_adjustments ADD COLUMN voided_by_adjustment_id INTEGER")
    if "correction_of_adjustment_id" not in columns:
        conn.execute("ALTER TABLE session_adjustments ADD COLUMN correction_of_adjustment_id INTEGER")
    if "correction_reason" not in columns:
        conn.execute("ALTER TABLE session_adjustments ADD COLUMN correction_reason TEXT")
    if "correction_note" not in columns:
        conn.execute("ALTER TABLE session_adjustments ADD COLUMN correction_note TEXT")


def row_to_patient_summary(row: sqlite3.Row) -> dict[str, Any]:
    return {
        "id": row["id"],
        "order": row["import_order"],
        "originalName": row["original_name"],
        "name": row["name"],
        "gender": row["gender"] or "",
        "age": row["age"] or "",
        "phone": row["phone"] or "",
        "address": row["address"] or "",
        "recordNo": row["record_no"],
        "remainingSessions": row["remaining_sessions"],
        "status": row["status"],
        "rechargeCount": row["recharge_count"] or 0,
        "updatedAt": row["updated_at"],
    }


def list_patients() -> list[dict[str, Any]]:
    with connect_db() as conn:
        rows = conn.execute(
            """
            SELECT p.*, COUNT(r.id) AS recharge_count
            FROM patients p
            LEFT JOIN recharges r ON r.patient_id = p.id
            GROUP BY p.id
            ORDER BY p.import_order ASC, p.id ASC
            """
        ).fetchall()
    return [row_to_patient_summary(row) for row in rows]


def create_patient(payload: dict[str, Any] | None = None) -> dict[str, Any]:
    payload = payload or {}
    name = str(payload.get("name") or "新建卡片").strip() or "新建卡片"
    created_at = now_text()
    with connect_db() as conn:
        next_order = conn.execute(
            "SELECT COALESCE(MAX(import_order), 0) + 1 AS next_order FROM patients"
        ).fetchone()["next_order"]
        cursor = conn.execute(
            """
            INSERT INTO patients (
                import_order, source_area, source_seq, original_name, name,
                status, created_at, updated_at
            )
            VALUES (?, ?, '', ?, ?, 'pending', ?, ?)
            """,
            (next_order, "手动新增", name, name, created_at, created_at),
        )
        conn.commit()
        patient_id = int(cursor.lastrowid)

    patient = get_patient(patient_id)
    if not patient:
        raise ValueError("新建卡片后读取失败。")
    return patient


def get_patient(patient_id: int) -> dict[str, Any] | None:
    with connect_db() as conn:
        patient = conn.execute("SELECT * FROM patients WHERE id = ?", (patient_id,)).fetchone()
        if not patient:
            return None
        recharges = conn.execute(
            """
            SELECT recharge_date, amount, sessions, raw_text
            FROM recharges
            WHERE patient_id = ?
            ORDER BY position ASC, id ASC
            """,
            (patient_id,),
        ).fetchall()

    return {
        "id": patient["id"],
        "order": patient["import_order"],
        "sourceArea": patient["source_area"] or "",
        "sourceSeq": patient["source_seq"] or "",
        "originalName": patient["original_name"],
        "name": patient["name"],
        "gender": patient["gender"] or "",
        "age": patient["age"] or "",
        "weight": patient["weight"] or "",
        "height": patient["height"] or "",
        "address": patient["address"] or "",
        "recordNo": patient["record_no"],
        "phone": patient["phone"] or "",
        "remainingSessions": patient["remaining_sessions"],
        "rawTranscript": patient["raw_transcript"] or "",
        "notes": patient["notes"] or "",
        "status": patient["status"],
        "createdAt": patient["created_at"],
        "updatedAt": patient["updated_at"],
        "recharges": [
            {
                "date": row["recharge_date"] or "",
                "amount": row["amount"],
                "sessions": row["sessions"],
                "rawText": row["raw_text"] or "",
            }
            for row in recharges
        ],
    }


def normalize_patient_address(value: Any) -> str:
    address = str(value or "").strip()
    if address and address not in PATIENT_ADDRESS_OPTIONS:
        raise ValueError("地址必须从固定选项中选择")
    return address


def normalize_patient_record_no(value: Any) -> int | None:
    if value in (None, ""):
        return None
    text = str(value).strip()
    if not re.fullmatch(r"\d+", text):
        raise ValueError("编号必须是正整数")
    record_no = int(text)
    if record_no <= 0:
        raise ValueError("编号必须是大于 0 的整数")
    return record_no


def save_patient(patient_id: int, payload: dict[str, Any]) -> dict[str, Any]:
    current = get_patient(patient_id)
    if not current:
        raise ValueError("没有找到这个人名。")

    status = str(payload.get("status") or "pending").strip()
    if status not in {"pending", "completed", "review"}:
        status = "pending"

    remaining = payload.get("remainingSessions")
    remaining_sessions = None
    if remaining not in (None, ""):
        remaining_sessions = int(remaining)

    address = normalize_patient_address(payload.get("address"))
    record_no = normalize_patient_record_no(
        payload.get("recordNo") if "recordNo" in payload else current.get("recordNo")
    )

    updated_at = now_text()
    with connect_db() as conn:
        duplicate = conn.execute(
            """
            SELECT id, name
            FROM patients
            WHERE id != ?
              AND COALESCE(address, '') = ?
              AND record_no = ?
            LIMIT 1
            """,
            (patient_id, address, record_no),
        ).fetchone() if record_no is not None else None
        if duplicate:
            address_label = address or "无地址"
            raise ValueError(
                f"{address_label}的编号 {record_no} 已被患者“{duplicate['name']}”使用"
            )
        conn.execute(
            """
            UPDATE patients
            SET name = ?, gender = ?, age = ?, weight = ?, height = ?, address = ?, record_no = ?,
                phone = ?, remaining_sessions = ?, raw_transcript = ?, notes = ?,
                status = ?, updated_at = ?
            WHERE id = ?
            """,
            (
                str(payload.get("name") or current["name"]).strip(),
                str(payload.get("gender") or "").strip(),
                str(payload.get("age") or "").strip(),
                str(payload.get("weight") or "").strip(),
                str(payload.get("height") or "").strip(),
                address,
                record_no,
                str(payload.get("phone") or "").strip(),
                remaining_sessions,
                str(payload.get("rawTranscript") or "").strip(),
                str(payload.get("notes") or "").strip(),
                status,
                updated_at,
                patient_id,
            ),
        )
        conn.execute("DELETE FROM recharges WHERE patient_id = ?", (patient_id,))
        for index, recharge in enumerate(payload.get("recharges") or [], start=1):
            if not isinstance(recharge, dict):
                continue
            amount = optional_int(recharge.get("amount"))
            sessions = optional_int(recharge.get("sessions"))
            recharge_date = str(recharge.get("date") or "").strip()
            raw_text = str(recharge.get("rawText") or "").strip()
            if not (recharge_date or amount is not None or sessions is not None or raw_text):
                continue
            conn.execute(
                """
                INSERT INTO recharges (
                    patient_id, position, recharge_date, amount, sessions, raw_text, created_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (patient_id, index, recharge_date, amount, sessions, raw_text, updated_at),
            )
        conn.commit()

    saved = get_patient(patient_id)
    if not saved:
        raise ValueError("保存后读取失败。")
    return saved


def row_to_session_adjustment(row: sqlite3.Row) -> dict[str, Any]:
    return {
        "id": row["id"],
        "patientId": row["patient_id"],
        "operation": row["operation"],
        "sessions": row["sessions"],
        "amount": row["amount"],
        "therapist": row["therapist"] or "",
        "beforeSessions": row["before_sessions"],
        "afterSessions": row["after_sessions"],
        "signatureStatus": row["signature_status"] or "pending",
        "signatureUrl": row["signature_url"] or "",
        "signatureSavedAt": row["signature_saved_at"] or "",
        "signatureSigner": row["signature_signer"] or "",
        "signatureNote": row["signature_note"] or "",
        "signatureBatchId": row["signature_batch_id"] or "",
        "voidedAt": row["voided_at"] or "",
        "voidedByAdjustmentId": row["voided_by_adjustment_id"] or "",
        "correctionOfAdjustmentId": row["correction_of_adjustment_id"] or "",
        "correctionReason": row["correction_reason"] or "",
        "correctionNote": row["correction_note"] or "",
        "isVoided": bool(row["voided_at"]),
        "isCorrection": bool(row["correction_of_adjustment_id"]),
        "occurredAt": row["occurred_at"],
        "note": row["note"] or "",
        "createdAt": row["created_at"],
    }


def list_session_adjustments(patient_id: int) -> list[dict[str, Any]]:
    with connect_db() as conn:
        rows = conn.execute(
            """
            SELECT *
            FROM session_adjustments
            WHERE patient_id = ?
              AND correction_of_adjustment_id IS NULL
            ORDER BY occurred_at DESC, id DESC
            """,
            (patient_id,),
        ).fetchall()
    return [row_to_session_adjustment(row) for row in rows]


def build_session_summary(query: dict[str, list[str]]) -> dict[str, Any]:
    range_type = first_query_value(query, "range") or "day"
    if range_type not in {"day", "month", "year"}:
        raise ValueError("统计范围必须是 day、month 或 year")

    date_value = first_query_value(query, "date")
    if not date_value:
        now = dt.datetime.now()
        if range_type == "day":
            date_value = now.strftime("%Y-%m-%d")
        elif range_type == "month":
            date_value = now.strftime("%Y-%m")
        else:
            date_value = now.strftime("%Y")

    date_prefix = normalize_summary_date(range_type, date_value)
    prefix_length = len(date_prefix)
    patient_id = optional_int(first_query_value(query, "patientId"))
    therapist = first_query_value(query, "therapist")
    if therapist and therapist not in THERAPISTS:
        raise ValueError("师傅筛选无效")

    adjustment_clauses = [f"substr(a.occurred_at, 1, {prefix_length}) = ?"]
    adjustment_params: list[Any] = [date_prefix]
    if patient_id:
        adjustment_clauses.append("a.patient_id = ?")
        adjustment_params.append(patient_id)

    adjustment_where_sql = " AND ".join(adjustment_clauses)
    with connect_db() as conn:
        adjustment_rows = conn.execute(
            f"""
            SELECT
                a.*,
                p.name AS patient_name,
                p.import_order AS patient_order,
                p.address AS patient_address,
                p.record_no AS patient_record_no
            FROM session_adjustments a
            JOIN patients p ON p.id = a.patient_id
            WHERE {adjustment_where_sql}
            ORDER BY a.occurred_at DESC, a.id DESC
            """,
            adjustment_params,
        ).fetchall()
        recharge_clauses = [f"substr(r.recharge_date, 1, {prefix_length}) = ?"]
        recharge_params: list[Any] = [date_prefix]
        if patient_id:
            recharge_clauses.append("r.patient_id = ?")
            recharge_params.append(patient_id)
        recharge_where_sql = " AND ".join(recharge_clauses)
        legacy_recharge_rows = conn.execute(
            f"""
            SELECT
                r.*,
                p.name AS patient_name,
                p.import_order AS patient_order,
                p.address AS patient_address,
                p.record_no AS patient_record_no
            FROM recharges r
            JOIN patients p ON p.id = r.patient_id
            WHERE {recharge_where_sql}
            ORDER BY r.recharge_date DESC, r.id DESC
            """,
            recharge_params,
        ).fetchall()

    adjustment_records = [row_to_summary_adjustment(row) for row in adjustment_rows]
    legacy_recharge_records = [row_to_legacy_recharge_summary(row) for row in legacy_recharge_rows]
    visible_adjustment_records = [item for item in adjustment_records if not item.get("isCorrection")]
    all_increase_records = [item for item in visible_adjustment_records if item["operation"] == "increase"] + legacy_recharge_records
    all_decrease_records = [item for item in visible_adjustment_records if item["operation"] == "decrease"]
    effective_increase_records = [item for item in all_increase_records if not item.get("isVoided")]
    effective_decrease_records = [item for item in all_decrease_records if not item.get("isVoided")]
    decrease_records = [
        item for item in all_decrease_records if not therapist or item["therapist"] == therapist
    ]
    effective_decrease_records_for_filter = [
        item for item in decrease_records if not item.get("isVoided")
    ]
    records = all_increase_records + decrease_records
    records.sort(key=lambda item: (item["occurredAt"], item["id"]), reverse=True)
    therapist_stats = []
    stat_names = [therapist] if therapist else list(THERAPISTS)
    for name in stat_names:
        therapist_decreases = [
            item
            for item in effective_decrease_records
            if item["operation"] == "decrease" and item["therapist"] == name
        ]
        therapist_stats.append(
            {
                "therapist": name,
                "workSessions": sum_number(item["sessions"] for item in therapist_decreases),
                "workCount": signed_record_count(therapist_decreases),
            }
        )

    return {
        "filters": {
            "range": range_type,
            "date": date_prefix,
            "patientId": patient_id or "",
            "therapist": therapist or "",
        },
        "therapists": list(THERAPISTS),
        "summary": {
            "recordCount": len(records),
            "patientCount": len({item["patientId"] for item in records}),
            "rechargeCount": signed_record_count(effective_increase_records),
            "rechargeSessions": sum_number(item["sessions"] for item in effective_increase_records),
            "rechargeAmount": sum_number(item["amount"] for item in effective_increase_records),
            "massageCount": signed_record_count(effective_decrease_records_for_filter),
            "massageSessions": sum_number(item["sessions"] for item in effective_decrease_records_for_filter),
        },
        "therapistStats": therapist_stats,
        "debts": list_debtor_patients(),
        "records": records,
    }


def first_query_value(query: dict[str, list[str]], key: str) -> str:
    values = query.get(key) or []
    return str(values[0]).strip() if values else ""


def normalize_summary_date(range_type: str, value: str) -> str:
    value = str(value or "").strip()
    if range_type == "day":
        if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", value):
            raise ValueError("日期格式应为 YYYY-MM-DD")
        return value
    if range_type == "month":
        if not re.fullmatch(r"\d{4}-\d{2}", value):
            raise ValueError("月份格式应为 YYYY-MM")
        return value
    if not re.fullmatch(r"\d{4}", value):
        raise ValueError("年份格式应为 YYYY")
    return value


def row_to_summary_adjustment(row: sqlite3.Row) -> dict[str, Any]:
    item = row_to_session_adjustment(row)
    item["patientName"] = row["patient_name"]
    item["patientOrder"] = row["patient_order"]
    item["patientAddress"] = row["patient_address"] or ""
    item["patientRecordNo"] = row["patient_record_no"]
    return item


def row_to_legacy_recharge_summary(row: sqlite3.Row) -> dict[str, Any]:
    return {
        "id": row["id"],
        "patientId": row["patient_id"],
        "patientName": row["patient_name"],
        "patientOrder": row["patient_order"],
        "patientAddress": row["patient_address"] or "",
        "patientRecordNo": row["patient_record_no"],
        "operation": "legacy_recharge",
        "sessions": row["sessions"],
                "amount": row["amount"],
                "therapist": "",
                "beforeSessions": None,
                "afterSessions": None,
                "signatureStatus": "legacy",
                "signatureUrl": "",
        "signatureSavedAt": "",
        "signatureSigner": "",
        "signatureNote": "",
        "signatureBatchId": "",
        "voidedAt": "",
        "voidedByAdjustmentId": "",
        "correctionOfAdjustmentId": "",
        "correctionReason": "",
        "correctionNote": "",
        "isVoided": False,
        "isCorrection": False,
        "occurredAt": row["recharge_date"] or "",
        "note": row["raw_text"] or "original recharge record",
        "createdAt": row["created_at"],
    }


def list_debtor_patients() -> list[dict[str, Any]]:
    conn = connect_db()
    try:
        rows = conn.execute(
            """
            SELECT id, import_order, name, phone, address, record_no, remaining_sessions
            FROM patients
            WHERE remaining_sessions < 0
            ORDER BY remaining_sessions ASC, import_order ASC
            """
        ).fetchall()
        patient_ids = [int(row["id"]) for row in rows]
        adjustment_rows: list[sqlite3.Row] = []
        if patient_ids:
            placeholders = ",".join("?" for _ in patient_ids)
            adjustment_rows = conn.execute(
                f"""
                SELECT patient_id, operation, sessions, occurred_at
                FROM session_adjustments
                WHERE patient_id IN ({placeholders})
                  AND voided_at IS NULL
                  AND correction_of_adjustment_id IS NULL
                ORDER BY patient_id ASC, occurred_at ASC, id ASC
                """,
                patient_ids,
            ).fetchall()
    finally:
        conn.close()

    adjustments_by_patient: dict[int, list[sqlite3.Row]] = {}
    for adjustment in adjustment_rows:
        adjustments_by_patient.setdefault(int(adjustment["patient_id"]), []).append(adjustment)

    debts = []
    for row in rows:
        patient_id = int(row["id"])
        current_balance = int(row["remaining_sessions"] or 0)
        adjustments = adjustments_by_patient.get(patient_id, [])
        total_effect = sum(
            int(item["sessions"] or 0) * (-1 if item["operation"] == "decrease" else 1)
            for item in adjustments
        )
        balance = current_balance - total_effect
        debt_since = ""
        for item in adjustments:
            before_balance = balance
            sessions = int(item["sessions"] or 0)
            balance += -sessions if item["operation"] == "decrease" else sessions
            if before_balance >= 0 and balance < 0:
                debt_since = item["occurred_at"] or ""
            elif balance >= 0:
                debt_since = ""

        debts.append(
            {
                "patientId": patient_id,
                "patientOrder": row["import_order"],
                "patientName": row["name"],
                "patientAddress": row["address"] or "",
                "patientRecordNo": row["record_no"],
                "phone": row["phone"] or "",
                "debtSince": debt_since,
                "remainingSessions": current_balance,
                "owedSessions": abs(current_balance),
            }
        )
    return debts


def normalize_settlement_date(value: Any, field_name: str) -> str:
    text = str(value or "").strip()
    try:
        parsed = dt.datetime.strptime(text, "%Y-%m-%d").date()
    except ValueError as exc:
        raise ValueError(f"{field_name}格式应为 YYYY-MM-DD") from exc
    return parsed.isoformat()


def list_debtor_patients_at(end_date: str) -> list[dict[str, Any]]:
    conn = connect_db()
    try:
        patients = conn.execute(
            """
            SELECT id, import_order, name, phone, address, record_no, remaining_sessions
            FROM patients
            ORDER BY import_order ASC
            """
        ).fetchall()
        adjustment_rows = conn.execute(
            """
            SELECT patient_id, operation, sessions, occurred_at
            FROM session_adjustments
            WHERE voided_at IS NULL
              AND correction_of_adjustment_id IS NULL
            ORDER BY patient_id ASC, occurred_at ASC, id ASC
            """
        ).fetchall()
    finally:
        conn.close()

    adjustments_by_patient: dict[int, list[sqlite3.Row]] = {}
    for adjustment in adjustment_rows:
        adjustments_by_patient.setdefault(int(adjustment["patient_id"]), []).append(adjustment)

    debts: list[dict[str, Any]] = []
    for patient in patients:
        patient_id = int(patient["id"])
        adjustments = adjustments_by_patient.get(patient_id, [])
        current_balance = int(patient["remaining_sessions"] or 0)
        total_effect = sum(
            int(item["sessions"] or 0) * (-1 if item["operation"] == "decrease" else 1)
            for item in adjustments
        )
        balance = current_balance - total_effect
        debt_since = ""
        for item in adjustments:
            if str(item["occurred_at"] or "")[:10] > end_date:
                break
            before_balance = balance
            sessions = int(item["sessions"] or 0)
            balance += -sessions if item["operation"] == "decrease" else sessions
            if before_balance >= 0 and balance < 0:
                debt_since = item["occurred_at"] or ""
            elif balance >= 0:
                debt_since = ""

        if balance < 0:
            debts.append(
                {
                    "patientId": patient_id,
                    "patientOrder": patient["import_order"],
                    "patientName": patient["name"],
                    "patientAddress": patient["address"] or "",
                    "patientRecordNo": patient["record_no"],
                    "phone": patient["phone"] or "",
                    "debtSince": debt_since,
                    "remainingSessions": balance,
                    "owedSessions": abs(balance),
                }
            )

    debts.sort(key=lambda item: (-int(item["owedSessions"]), int(item["patientOrder"])))
    return debts


def build_settlement_snapshot(start_date: str, end_date: str) -> dict[str, Any]:
    conn = connect_db()
    try:
        adjustment_rows = conn.execute(
            """
            SELECT a.*, p.name AS patient_name, p.import_order AS patient_order,
                   p.address AS patient_address, p.record_no AS patient_record_no
            FROM session_adjustments a
            JOIN patients p ON p.id = a.patient_id
            WHERE substr(a.occurred_at, 1, 10) BETWEEN ? AND ?
            ORDER BY a.occurred_at DESC, a.id DESC
            """,
            (start_date, end_date),
        ).fetchall()
        legacy_rows = conn.execute(
            """
            SELECT r.*, p.name AS patient_name, p.import_order AS patient_order,
                   p.address AS patient_address, p.record_no AS patient_record_no
            FROM recharges r
            JOIN patients p ON p.id = r.patient_id
            WHERE substr(r.recharge_date, 1, 10) BETWEEN ? AND ?
            ORDER BY r.recharge_date DESC, r.id DESC
            """,
            (start_date, end_date),
        ).fetchall()
    finally:
        conn.close()

    adjustment_records = [row_to_summary_adjustment(row) for row in adjustment_rows]
    visible_adjustments = [item for item in adjustment_records if not item.get("isCorrection")]
    legacy_records = [row_to_legacy_recharge_summary(row) for row in legacy_rows]
    records = visible_adjustments + legacy_records
    records.sort(key=lambda item: (item["occurredAt"], item["id"]), reverse=True)

    effective_increases = [
        item
        for item in records
        if item["operation"] in {"increase", "legacy_recharge"} and not item.get("isVoided")
    ]
    effective_decreases = [
        item
        for item in records
        if item["operation"] == "decrease" and not item.get("isVoided")
    ]
    therapist_stats = []
    for name in THERAPISTS:
        therapist_records = [item for item in effective_decreases if item["therapist"] == name]
        therapist_stats.append(
            {
                "therapist": name,
                "workSessions": sum_number(item["sessions"] for item in therapist_records),
                "workCount": signed_record_count(therapist_records),
            }
        )

    return {
        "version": 1,
        "startDate": start_date,
        "endDate": end_date,
        "generatedAt": now_text(),
        "summary": {
            "recordCount": len(records),
            "patientCount": len({item["patientId"] for item in records}),
            "rechargeCount": signed_record_count(effective_increases),
            "rechargeSessions": sum_number(item["sessions"] for item in effective_increases),
            "rechargeAmount": sum_number(item["amount"] for item in effective_increases),
            "massageCount": signed_record_count(effective_decreases),
            "massageSessions": sum_number(item["sessions"] for item in effective_decreases),
        },
        "therapistStats": therapist_stats,
        "debts": list_debtor_patients_at(end_date),
        "records": records,
    }


def parse_settlement_snapshot(row: sqlite3.Row) -> dict[str, Any]:
    try:
        snapshot = json.loads(row["snapshot_json"])
    except (TypeError, json.JSONDecodeError) as exc:
        raise ValueError("月结快照损坏，无法读取") from exc
    if not isinstance(snapshot, dict):
        raise ValueError("月结快照格式错误")
    return snapshot


def row_to_settlement(row: sqlite3.Row, latest_active_id: int | None, include_snapshot: bool) -> dict[str, Any]:
    snapshot = parse_settlement_snapshot(row)
    summary = snapshot.get("summary") if isinstance(snapshot.get("summary"), dict) else {}
    item = {
        "id": int(row["id"]),
        "startDate": row["start_date"],
        "endDate": row["end_date"],
        "status": row["status"],
        "createdAt": row["created_at"],
        "revokedAt": row["revoked_at"] or "",
        "revokedReason": row["revoked_reason"] or "",
        "canRevoke": row["status"] == "active" and int(row["id"]) == latest_active_id,
        "summary": summary,
    }
    if include_snapshot:
        item["snapshot"] = snapshot
    return item


def latest_active_settlement_row(conn: sqlite3.Connection) -> sqlite3.Row | None:
    return conn.execute(
        """
        SELECT *
        FROM settlements
        WHERE status = 'active'
        ORDER BY end_date DESC, id DESC
        LIMIT 1
        """
    ).fetchone()


def settlement_default_start(latest: sqlite3.Row | None) -> str:
    if not latest:
        return dt.date.today().isoformat()
    end_date = dt.datetime.strptime(latest["end_date"], "%Y-%m-%d").date()
    return (end_date + dt.timedelta(days=1)).isoformat()


def list_settlements() -> dict[str, Any]:
    conn = connect_db()
    try:
        latest = latest_active_settlement_row(conn)
        latest_id = int(latest["id"]) if latest else None
        rows = conn.execute(
            """
            SELECT *
            FROM settlements
            WHERE status = 'active'
            ORDER BY end_date DESC, id DESC
            """
        ).fetchall()
        revoked_count = int(
            conn.execute("SELECT COUNT(*) AS count FROM settlements WHERE status = 'revoked'").fetchone()["count"]
        )
    finally:
        conn.close()

    today = dt.date.today().isoformat()
    default_start = settlement_default_start(latest)
    return {
        "settlements": [row_to_settlement(row, latest_id, False) for row in rows],
        "defaults": {
            "startDate": default_start,
            "endDate": today,
            "startLocked": bool(latest),
            "canCreate": default_start <= today,
        },
        "revokedCount": revoked_count,
    }


def get_settlement(settlement_id: int) -> dict[str, Any] | None:
    conn = connect_db()
    try:
        latest = latest_active_settlement_row(conn)
        latest_id = int(latest["id"]) if latest else None
        row = conn.execute("SELECT * FROM settlements WHERE id = ?", (settlement_id,)).fetchone()
    finally:
        conn.close()
    return row_to_settlement(row, latest_id, True) if row else None


def active_settlement_for_date(conn: sqlite3.Connection, date_value: str) -> sqlite3.Row | None:
    return conn.execute(
        """
        SELECT id, start_date, end_date
        FROM settlements
        WHERE status = 'active'
          AND ? BETWEEN start_date AND end_date
        ORDER BY end_date DESC, id DESC
        LIMIT 1
        """,
        (date_value,),
    ).fetchone()


def create_settlement(payload: dict[str, Any]) -> dict[str, Any]:
    start_date = normalize_settlement_date(payload.get("startDate"), "起始日期")
    end_date = normalize_settlement_date(payload.get("endDate"), "结束日期")
    if start_date > end_date:
        raise ValueError("起始日期不能晚于结束日期")
    if end_date > dt.date.today().isoformat():
        raise ValueError("结束日期不能晚于今天")

    conn = connect_db()
    try:
        latest = latest_active_settlement_row(conn)
        expected_start = settlement_default_start(latest)
        if latest and start_date != expected_start:
            raise ValueError(f"本次起始日期必须是上一份月结结束后的下一天：{expected_start}")
        overlap = conn.execute(
            """
            SELECT id, start_date, end_date
            FROM settlements
            WHERE status = 'active'
              AND NOT (end_date < ? OR start_date > ?)
            LIMIT 1
            """,
            (start_date, end_date),
        ).fetchone()
        if overlap:
            raise ValueError(
                f"所选日期与月结 #{overlap['id']}（{overlap['start_date']} 至 {overlap['end_date']}）重叠"
            )
    finally:
        conn.close()

    snapshot = build_settlement_snapshot(start_date, end_date)
    created_at = now_text()
    conn = connect_db()
    try:
        conn.execute("BEGIN IMMEDIATE")
        latest = latest_active_settlement_row(conn)
        expected_start = settlement_default_start(latest)
        if latest and start_date != expected_start:
            raise ValueError("月结状态已变化，请刷新页面后重试")
        cursor = conn.execute(
            """
            INSERT INTO settlements (start_date, end_date, status, snapshot_json, created_at)
            VALUES (?, ?, 'active', ?, ?)
            """,
            (start_date, end_date, json.dumps(snapshot, ensure_ascii=False), created_at),
        )
        settlement_id = int(cursor.lastrowid)
        conn.commit()
    finally:
        conn.close()

    settlement = get_settlement(settlement_id)
    if not settlement:
        raise ValueError("月结生成后读取失败")
    return settlement


def revoke_settlement(settlement_id: int, payload: dict[str, Any]) -> dict[str, Any]:
    reason = str(payload.get("reason") or "").strip()
    if not reason:
        raise ValueError("请填写撤销原因")
    if len(reason) > 200:
        raise ValueError("撤销原因不能超过 200 个字")

    revoked_at = now_text()
    conn = connect_db()
    try:
        conn.execute("BEGIN IMMEDIATE")
        latest = latest_active_settlement_row(conn)
        if not latest or int(latest["id"]) != settlement_id:
            raise ValueError("只能撤销最后一份有效月结")
        conn.execute(
            """
            UPDATE settlements
            SET status = 'revoked', revoked_at = ?, revoked_reason = ?
            WHERE id = ?
            """,
            (revoked_at, reason, settlement_id),
        )
        conn.commit()
    finally:
        conn.close()

    return {
        "revokedId": settlement_id,
        "revokedAt": revoked_at,
        "reason": reason,
        **list_settlements(),
    }


def sum_number(values: Any) -> float:
    total = 0.0
    for value in values:
        if value in (None, ""):
            continue
        total += float(value)
    return total


def signed_record_count(items: list[dict[str, Any]]) -> int:
    total = 0
    for item in items:
        if item.get("isVoided") or item.get("isCorrection"):
            continue
        if item.get("operation") == "legacy_recharge":
            total += 1
            continue
        try:
            sessions = float(item.get("sessions") or 0)
        except (TypeError, ValueError):
            sessions = 0
        if sessions > 0:
            total += 1
    return total


def backup_now() -> dt.datetime:
    return dt.datetime.now()


def backup_today_text() -> str:
    return backup_now().strftime("%Y-%m-%d")


def backup_iso(value: dt.datetime | None = None) -> str:
    return (value or backup_now()).replace(microsecond=0).isoformat(sep=" ")


def set_backup_status(state: str, message: str, **extra: Any) -> None:
    with BACKUP_STATUS_LOCK:
        BACKUP_STATUS.update({"state": state, "message": message, "updatedAt": backup_iso()})
        BACKUP_STATUS.update(extra)


def get_backup_status() -> dict[str, Any]:
    with BACKUP_STATUS_LOCK:
        status = dict(BACKUP_STATUS)
    saved_state = read_backup_state()
    if saved_state:
        status.setdefault("lastSuccessAt", saved_state.get("lastSuccessAt", ""))
        status.setdefault("lastRemoteName", saved_state.get("lastRemoteName", ""))
        status.setdefault("lastIncludedSignatures", saved_state.get("lastIncludedSignatures", 0))
        status.setdefault("lastPackageBytes", saved_state.get("lastPackageBytes", 0))
    return status


def read_backup_state() -> dict[str, Any]:
    if not BACKUP_STATE_PATH.exists():
        return {}
    try:
        data = json.loads(BACKUP_STATE_PATH.read_text(encoding="utf-8"))
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def write_backup_state(data: dict[str, Any]) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    BACKUP_STATE_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def backup_config() -> dict[str, str]:
    webdav_url = os.environ.get("BACKUP_WEBDAV_URL", "").strip()
    username = os.environ.get("BACKUP_WEBDAV_USERNAME", "").strip()
    password = os.environ.get("BACKUP_WEBDAV_PASSWORD", "").strip()
    encryption_password = os.environ.get("BACKUP_ENCRYPTION_PASSWORD", "").strip()
    app_folder = os.environ.get("BACKUP_WEBDAV_APP", "").strip().strip("/")

    if webdav_url and app_folder:
        parsed = urllib.parse.urlparse(webdav_url)
        current_path = parsed.path.rstrip("/")
        if not current_path.endswith("/" + app_folder):
            webdav_url = urllib.parse.urlunparse(parsed._replace(path=f"{current_path}/{app_folder}/"))
    if webdav_url and not webdav_url.endswith("/"):
        webdav_url += "/"

    missing = [
        name
        for name, value in (
            ("BACKUP_WEBDAV_URL", webdav_url),
            ("BACKUP_WEBDAV_USERNAME", username),
            ("BACKUP_WEBDAV_PASSWORD", password),
            ("BACKUP_ENCRYPTION_PASSWORD", encryption_password),
        )
        if not value
    ]
    if missing:
        raise ValueError(f"云备份未配置：缺少 {', '.join(missing)}")
    return {
        "url": webdav_url,
        "username": username,
        "password": password,
        "encryptionPassword": encryption_password,
        "appFolder": app_folder or APP_DIR.name,
    }


def backup_auth_header(config: dict[str, str]) -> str:
    token = base64.b64encode(f"{config['username']}:{config['password']}".encode("utf-8")).decode("ascii")
    return f"Basic {token}"


def webdav_request(
    config: dict[str, str],
    method: str,
    url: str,
    data: bytes | None = None,
    headers: dict[str, str] | None = None,
    timeout: int = 30,
) -> bytes:
    request_headers = {
        "Authorization": backup_auth_header(config),
        "User-Agent": "tuina-local-backup/1.0",
    }
    if headers:
        request_headers.update(headers)
    request = urllib.request.Request(url, data=data, headers=request_headers, method=method)
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            return response.read()
    except urllib.error.HTTPError as exc:
        if method == "MKCOL" and exc.code in {HTTPStatus.METHOD_NOT_ALLOWED, HTTPStatus.CONFLICT}:
            return b""
        if method == "DELETE" and exc.code == HTTPStatus.NOT_FOUND:
            return b""
        raise


def ensure_webdav_collection(config: dict[str, str]) -> None:
    webdav_request(config, "MKCOL", config["url"], timeout=20)


def create_database_snapshot(snapshot_path: Path) -> None:
    snapshot_path.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(DB_PATH) as source, sqlite3.connect(snapshot_path) as target:
        source.backup(target)


def derive_backup_key(password: str, salt: bytes) -> bytes:
    key = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, 390000, dklen=32)
    return base64.urlsafe_b64encode(key)


def encrypt_backup_file(source_path: Path, encrypted_path: Path, password: str) -> None:
    from cryptography.fernet import Fernet

    salt = os.urandom(16)
    encrypted = Fernet(derive_backup_key(password, salt)).encrypt(source_path.read_bytes())
    encrypted_path.write_bytes(BACKUP_MAGIC + salt + encrypted)


def backup_filename(now: dt.datetime, app_name: str) -> str:
    safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", app_name).strip("_") or "tuina"
    return f"{safe_name}_backup_{now.strftime('%Y-%m-%d_%H%M%S')}.zip.enc"


def signature_backup_record(path: Path) -> dict[str, Any]:
    stat = path.stat()
    return {"mtimeNs": stat.st_mtime_ns, "size": stat.st_size}


def backup_archive_name(path: Path) -> str:
    try:
        return path.relative_to(APP_DIR).as_posix()
    except ValueError:
        pass
    try:
        return "doc_patient/signature_patient/" + path.relative_to(SIGNATURE_PATIENT_DIR).as_posix()
    except ValueError:
        return path.name


def collect_incremental_signature_paths(state: dict[str, Any]) -> list[Path]:
    uploaded = state.get("uploadedSignatureFiles")
    if not isinstance(uploaded, dict):
        uploaded = {}
    if not ELECTRONIC_SIGNATURE_DIR.exists():
        return []
    paths: list[Path] = []
    for path in sorted(ELECTRONIC_SIGNATURE_DIR.rglob("*.png")):
        if not path.is_file():
            continue
        rel = backup_archive_name(path)
        if uploaded.get(rel) != signature_backup_record(path):
            paths.append(path)
    return paths


def add_file_to_zip(package: zipfile.ZipFile, path: Path, arcname: str) -> None:
    if path.exists() and path.is_file():
        package.write(path, arcname)


def create_backup_zip(zip_path: Path, state: dict[str, Any]) -> tuple[list[Path], dict[str, Any]]:
    BACKUP_TEMP_DIR.mkdir(parents=True, exist_ok=True)
    snapshot_path = BACKUP_TEMP_DIR / "tuina_records_snapshot.sqlite3"
    create_database_snapshot(snapshot_path)
    signature_paths = collect_incremental_signature_paths(state)
    uploaded = dict(state.get("uploadedSignatureFiles") or {})
    manifest = {
        "generatedAt": backup_iso(),
        "database": "data/tuina_records.sqlite3",
        "includedSignatureCount": len(signature_paths),
        "includedSignatures": [backup_archive_name(path) for path in signature_paths],
    }

    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as package:
        package.write(snapshot_path, "data/tuina_records.sqlite3")
        add_file_to_zip(package, SIGNATURE_BINDINGS_PATH, "doc_patient/signature_patient/bindings.json")
        add_file_to_zip(package, SIGNATURE_MANIFEST_PATH, "doc_patient/signature_patient/manifest.json")
        for log_path in sorted(DATA_DIR.glob("*.log")):
            if log_path.is_file() and log_path.stat().st_size <= 1024 * 1024:
                package.write(log_path, log_path.relative_to(APP_DIR).as_posix())
        for signature_path in signature_paths:
            arcname = backup_archive_name(signature_path)
            package.write(signature_path, arcname)
            uploaded[arcname] = signature_backup_record(signature_path)
        package.writestr("backup_manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))

    updated_state = dict(state)
    updated_state["uploadedSignatureFiles"] = uploaded
    return signature_paths, updated_state


def upload_backup_file(config: dict[str, str], encrypted_path: Path, remote_name: str) -> None:
    remote_url = urllib.parse.urljoin(config["url"], urllib.parse.quote(remote_name))
    webdav_request(
        config,
        "PUT",
        remote_url,
        data=encrypted_path.read_bytes(),
        headers={"Content-Type": "application/octet-stream"},
        timeout=60,
    )


def webdav_backup_names(config: dict[str, str]) -> list[str]:
    body = webdav_request(
        config,
        "PROPFIND",
        config["url"],
        headers={"Depth": "1", "Content-Type": "application/xml"},
        timeout=30,
    )
    names: list[str] = []
    try:
        root = ET.fromstring(body)
    except ET.ParseError:
        return names
    namespace = {"d": "DAV:"}
    for response in root.findall("d:response", namespace):
        href = response.findtext("d:href", default="", namespaces=namespace)
        name = urllib.parse.unquote(href.rstrip("/").rsplit("/", 1)[-1])
        if name.endswith(".zip.enc") and "_backup_" in name:
            names.append(name)
    return names


def backup_datetime_from_name(name: str) -> dt.datetime | None:
    match = re.search(r"_backup_(\d{4}-\d{2}-\d{2}_\d{6})\.zip\.enc$", name)
    if not match:
        return None
    try:
        return dt.datetime.strptime(match.group(1), "%Y-%m-%d_%H%M%S")
    except ValueError:
        return None


def cleanup_remote_backups(config: dict[str, str]) -> None:
    backups: list[tuple[str, dt.datetime]] = []
    for name in webdav_backup_names(config):
        backup_time = backup_datetime_from_name(name)
        if backup_time:
            backups.append((name, backup_time))
    cutoff = backup_now() - dt.timedelta(days=BACKUP_RETENTION_DAYS)
    backups.sort(key=lambda item: item[1], reverse=True)
    names_to_delete = {
        name
        for index, (name, backup_time) in enumerate(backups)
        if backup_time < cutoff or index >= BACKUP_RETENTION_DAYS
    }
    for name in sorted(names_to_delete):
        remote_url = urllib.parse.urljoin(config["url"], urllib.parse.quote(name))
        webdav_request(config, "DELETE", remote_url, timeout=30)


def perform_cloud_backup(force: bool = False) -> dict[str, Any]:
    if not BACKUP_RUN_LOCK.acquire(blocking=False):
        raise ValueError("云备份正在运行")
    try:
        config = backup_config()
        saved_state = read_backup_state()
        today = backup_today_text()
        if not force and saved_state.get("lastSuccessDate") == today:
            return saved_state

        now = backup_now()
        remote_name = backup_filename(now, config["appFolder"])
        zip_path = BACKUP_TEMP_DIR / remote_name.replace(".zip.enc", ".zip")
        encrypted_path = BACKUP_TEMP_DIR / remote_name
        BACKUP_TEMP_DIR.mkdir(parents=True, exist_ok=True)
        try:
            signature_paths, updated_state = create_backup_zip(zip_path, saved_state)
            encrypt_backup_file(zip_path, encrypted_path, config["encryptionPassword"])
            ensure_webdav_collection(config)
            upload_backup_file(config, encrypted_path, remote_name)
            try:
                cleanup_remote_backups(config)
            except Exception as exc:
                print(f"[backup] cleanup failed: {exc}", file=sys.stderr, flush=True)

            result = {
                **updated_state,
                "lastSuccessDate": today,
                "lastSuccessAt": backup_iso(now),
                "lastRemoteName": remote_name,
                "lastIncludedSignatures": len(signature_paths),
                "lastPackageBytes": encrypted_path.stat().st_size,
            }
            write_backup_state(result)
            return result
        finally:
            for path in (zip_path, encrypted_path, BACKUP_TEMP_DIR / "tuina_records_snapshot.sqlite3"):
                try:
                    if path.exists():
                        path.unlink()
                except OSError:
                    pass
    finally:
        BACKUP_RUN_LOCK.release()


def backup_error_message(error: Exception) -> str:
    if isinstance(error, urllib.error.URLError):
        return "云备份失败：网络不可用，稍后自动重试"
    if isinstance(error, TimeoutError):
        return "云备份失败：网络超时，稍后自动重试"
    message = str(error)
    return f"云备份失败：{message or '稍后自动重试'}"


def cloud_backup_worker() -> None:
    try:
        backup_config()
    except Exception as exc:
        set_backup_status("disabled", str(exc), configured=False)
        return

    saved_state = read_backup_state()
    if saved_state.get("lastSuccessDate") == backup_today_text():
        set_backup_status(
            "success",
            "今日云备份已保存",
            configured=True,
            lastSuccessAt=saved_state.get("lastSuccessAt", ""),
            lastRemoteName=saved_state.get("lastRemoteName", ""),
            lastIncludedSignatures=saved_state.get("lastIncludedSignatures", 0),
            lastPackageBytes=saved_state.get("lastPackageBytes", 0),
        )
        return

    while True:
        set_backup_status("running", "云备份中", configured=True, nextRetryAt="")
        try:
            result = perform_cloud_backup(force=False)
            set_backup_status(
                "success",
                "今日云备份已保存",
                configured=True,
                lastSuccessAt=result.get("lastSuccessAt", ""),
                lastRemoteName=result.get("lastRemoteName", ""),
                lastIncludedSignatures=result.get("lastIncludedSignatures", 0),
                lastPackageBytes=result.get("lastPackageBytes", 0),
            )
            return
        except Exception as exc:
            next_retry = backup_now() + dt.timedelta(seconds=BACKUP_RETRY_SECONDS)
            set_backup_status(
                "error",
                backup_error_message(exc),
                configured=True,
                nextRetryAt=backup_iso(next_retry),
            )
            time.sleep(BACKUP_RETRY_SECONDS)


def start_cloud_backup_worker() -> None:
    global BACKUP_WORKER_STARTED
    if BACKUP_WORKER_STARTED:
        return
    BACKUP_WORKER_STARTED = True
    thread = threading.Thread(target=cloud_backup_worker, name="tuina-cloud-backup", daemon=True)
    thread.start()


def get_signature_item(patient_id: int) -> dict[str, Any]:
    if not SIGNATURE_MANIFEST_PATH.exists():
        return {}
    manifest = load_signature_manifest()
    for item in manifest.get("items") or []:
        if int(item.get("patientId") or 0) == patient_id:
            return item
    return {}


def get_visit_signature_history(patient_id: int) -> list[dict[str, Any]]:
    try:
        bindings = load_signature_bindings()
    except Exception:
        bindings = {"patients": {}}
    patient_binding = (bindings.get("patients") or {}).get(str(patient_id))
    items: list[dict[str, Any]] = []
    seen_urls: set[str] = set()
    history = patient_binding.get("history") if isinstance(patient_binding, dict) else []
    if isinstance(history, list):
        for entry in history:
            if not isinstance(entry, dict):
                continue
            if entry.get("kind") != "visit" and entry.get("field") != "visitSignature":
                continue
            url = str(entry.get("url") or "").strip()
            if not url:
                continue
            seen_urls.add(url)
            items.append(
                {
                    "url": url,
                    "savedAt": str(entry.get("savedAt") or "").strip(),
                    "signerName": str(entry.get("signerName") or "").strip(),
                    "note": str(entry.get("note") or "").strip(),
                    "adjustmentId": entry.get("adjustmentId") or "",
                }
            )

    with connect_db() as conn:
        rows = conn.execute(
            """
            SELECT id, signature_url, signature_saved_at, signature_signer, signature_note, occurred_at, sessions, therapist
            FROM session_adjustments
            WHERE patient_id = ?
              AND operation = 'decrease'
              AND signature_url IS NOT NULL
              AND signature_url != ''
              AND voided_at IS NULL
              AND correction_of_adjustment_id IS NULL
            ORDER BY occurred_at DESC, id DESC
            """,
            (patient_id,),
        ).fetchall()
    for row in rows:
        url = str(row["signature_url"] or "").strip()
        if not url:
            continue
        if url in seen_urls:
            continue
        seen_urls.add(url)
        items.append(
            {
                "url": url,
                "savedAt": str(row["signature_saved_at"] or row["occurred_at"] or "").strip(),
                "signerName": str(row["signature_signer"] or "").strip(),
                "note": str(row["signature_note"] or f"消费{row['sessions']}次，师傅{row['therapist'] or '未记录'}").strip(),
                "adjustmentId": row["id"],
            }
        )
    return items


def get_session_page(patient_id: int) -> dict[str, Any]:
    patient = get_patient(patient_id)
    if not patient:
        raise ValueError("patient not found")
    return {
        "patient": patient,
        "signature": get_signature_item(patient_id),
        "visitSignatures": get_visit_signature_history(patient_id),
        "adjustments": list_session_adjustments(patient_id),
    }


def apply_session_adjustment(patient_id: int, payload: dict[str, Any]) -> dict[str, Any]:
    operation = str(payload.get("operation") or "").strip()
    if operation not in {"increase", "decrease"}:
        raise ValueError("operation must be increase or decrease")
    sessions = optional_int(payload.get("sessions"))
    if sessions is None or sessions <= 0:
        raise ValueError("sessions must be a positive integer")
    therapist = str(payload.get("therapist") or "").strip() if operation == "decrease" else ""
    if operation == "decrease" and therapist not in THERAPISTS:
        raise ValueError("请选择师傅")
    amount = optional_float(payload.get("amount")) if operation == "increase" else None
    if operation == "increase":
        if amount is None:
            amount = float(sessions * DEFAULT_SESSION_PRICE)
        if amount < 0:
            raise ValueError("充值金额不能小于 0")

    occurred_at = str(payload.get("occurredAt") or "").strip() or now_text()
    note = str(payload.get("note") or "").strip()
    created_at = now_text()

    with connect_db() as conn:
        locked_settlement = active_settlement_for_date(conn, occurred_at[:10])
        if locked_settlement:
            raise ValueError(
                f"该日期已进入月结 #{locked_settlement['id']}（{locked_settlement['start_date']} 至 {locked_settlement['end_date']}），请先撤销对应月结"
            )
        patient = conn.execute(
            "SELECT id, remaining_sessions FROM patients WHERE id = ?",
            (patient_id,),
        ).fetchone()
        if not patient:
            raise ValueError("patient not found")

        before_sessions = patient["remaining_sessions"]
        before_value = int(before_sessions) if before_sessions is not None else 0
        if operation == "increase":
            after_sessions = before_value + sessions
        else:
            after_sessions = before_value - sessions

        conn.execute(
            """
            UPDATE patients
            SET remaining_sessions = ?, updated_at = ?
            WHERE id = ?
            """,
            (after_sessions, created_at, patient_id),
        )
        cursor = conn.execute(
            """
            INSERT INTO session_adjustments (
                patient_id, operation, sessions, amount, therapist, before_sessions, after_sessions, signature_status,
                occurred_at, note, created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                patient_id,
                operation,
                sessions,
                amount,
                therapist,
                before_sessions,
                after_sessions,
                "pending",
                occurred_at,
                note,
                created_at,
            ),
        )
        conn.commit()
        adjustment_id = int(cursor.lastrowid)

    return {
        "adjustment": {
            "id": adjustment_id,
            "patientId": patient_id,
            "operation": operation,
            "sessions": sessions,
            "amount": amount,
            "therapist": therapist,
            "beforeSessions": before_sessions,
            "afterSessions": after_sessions,
            "signatureStatus": "pending",
            "signatureUrl": "",
            "signatureSavedAt": "",
            "signatureSigner": "",
            "signatureNote": "",
            "occurredAt": occurred_at,
            "note": note,
            "createdAt": created_at,
        },
        "patient": get_patient(patient_id),
        "adjustments": list_session_adjustments(patient_id),
    }


def reverse_session_adjustment(adjustment_id: int, payload: dict[str, Any]) -> dict[str, Any]:
    reason = str(payload.get("reason") or "").strip()
    if reason not in CORRECTION_REASONS:
        raise ValueError("冲正原因只能选择：手误、患者反悔、其他")
    note_detail = str(payload.get("note") or "").strip()
    if reason == "其他" and not note_detail:
        raise ValueError("选择其他时请填写说明")

    corrected_at = now_text()
    with connect_db() as conn:
        original = conn.execute(
            """
            SELECT a.*, p.remaining_sessions
            FROM session_adjustments a
            JOIN patients p ON p.id = a.patient_id
            WHERE a.id = ?
            """,
            (adjustment_id,),
        ).fetchone()
        if not original:
            raise ValueError("没有找到要冲正的流水")
        if original["correction_of_adjustment_id"]:
            raise ValueError("冲正流水不能再次冲正")
        if original["voided_at"]:
            raise ValueError("这条流水已经被冲正")
        if original["operation"] not in {"increase", "decrease"}:
            raise ValueError("这条流水类型不能冲正")
        locked_settlement = active_settlement_for_date(conn, str(original["occurred_at"] or "")[:10])
        if locked_settlement:
            raise ValueError(
                f"这条流水已进入月结 #{locked_settlement['id']}（{locked_settlement['start_date']} 至 {locked_settlement['end_date']}），请先按顺序撤销月结"
            )

        patient_id = int(original["patient_id"])
        operation = str(original["operation"])
        original_sessions = int(original["sessions"])
        if original_sessions <= 0:
            raise ValueError("这条流水次数异常，不能冲正")
        before_sessions = int(original["remaining_sessions"] or 0)
        if operation == "increase":
            after_sessions = before_sessions - original_sessions
        else:
            after_sessions = before_sessions + original_sessions

        reason_text = reason if reason != "其他" else f"其他：{note_detail}"
        conn.execute(
            """
            UPDATE session_adjustments
            SET voided_at = ?,
                voided_by_adjustment_id = NULL,
                correction_reason = ?,
                correction_note = ?
            WHERE id = ?
            """,
            (corrected_at, reason, note_detail, adjustment_id),
        )
        conn.execute(
            """
            UPDATE patients
            SET remaining_sessions = ?, updated_at = ?
            WHERE id = ?
            """,
            (after_sessions, corrected_at, patient_id),
        )
        conn.commit()

    return {
        "patient": get_patient(patient_id),
        "reason": reason_text,
        "original": get_session_adjustment(adjustment_id),
    }


def get_session_adjustment(adjustment_id: int) -> dict[str, Any]:
    with connect_db() as conn:
        row = conn.execute("SELECT * FROM session_adjustments WHERE id = ?", (adjustment_id,)).fetchone()
    if not row:
        raise ValueError("没有找到流水")
    return row_to_session_adjustment(row)


def optional_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(float(str(value).strip()))
    except ValueError:
        return None


def optional_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(str(value).strip())
    except ValueError as exc:
        raise ValueError(f"invalid number: {value}") from exc


def export_csv() -> bytes:
    patients = list_patients()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(
        [
            "顺序",
            "原姓名",
            "当前姓名",
            "性别",
            "年龄",
            "电话",
            "体重",
            "身高",
            "地址",
            "地址编号",
            "充值记录",
            "剩余次数",
            "状态",
            "备注",
            "识别原文",
            "更新时间",
        ]
    )
    for summary in patients:
        patient = get_patient(int(summary["id"]))
        if not patient:
            continue
        recharge_text = "；".join(
            [
                f"{item.get('date') or '未填日期'} {item.get('amount') or ''}元 {item.get('sessions') or ''}次".strip()
                for item in patient["recharges"]
            ]
        )
        writer.writerow(
            [
                patient["order"],
                patient["originalName"],
                patient["name"],
                patient["gender"],
                patient["age"],
                patient["phone"],
                patient["weight"],
                patient["height"],
                patient["address"],
                patient["recordNo"] if patient["recordNo"] is not None else "",
                recharge_text,
                patient["remainingSessions"] if patient["remainingSessions"] is not None else "",
                patient["status"],
                patient["notes"],
                patient["rawTranscript"],
                patient["updatedAt"],
            ]
        )
    return ("\ufeff" + output.getvalue()).encode("utf-8-sig")


def public_config() -> dict[str, str]:
    return {
        "engine": get_env("TENCENT_ASR_ENGINE", default="16k_zh_medical"),
        "region": tencent_region(),
        "transport": "http",
        "hasCredentials": str(bool(tencent_secret_id() and tencent_secret_key())).lower(),
        "hotwordId": "configured" if os.getenv("TENCENT_ASR_HOTWORD_ID") else "",
    }


def get_env(*names: str, default: str = "") -> str:
    for name in names:
        value = os.getenv(name)
        if value:
            return value
    return default


def tencent_secret_id() -> str:
    return get_env("TENCENT_SECRET_ID", "TENCENTCLOUD_SECRET_ID")


def tencent_secret_key() -> str:
    return get_env("TENCENT_SECRET_KEY", "TENCENTCLOUD_SECRET_KEY")


def tencent_token() -> str:
    return get_env("TENCENT_TOKEN", "TENCENTCLOUD_TOKEN")


def tencent_region() -> str:
    return get_env("TENCENT_REGION", "TENCENTCLOUD_REGION", default="ap-guangzhou")


def hmac_sha256(key: bytes, message: str) -> bytes:
    return hmac.new(key, message.encode("utf-8"), hashlib.sha256).digest()


def sha256_hex(message: bytes | str) -> str:
    if isinstance(message, str):
        message = message.encode("utf-8")
    return hashlib.sha256(message).hexdigest()


def build_tencent_headers(payload: bytes) -> dict[str, str]:
    secret_id = tencent_secret_id()
    secret_key = tencent_secret_key()
    if not secret_id or not secret_key:
        raise RuntimeError("缺少腾讯云密钥，请检查 tecent_api_key.txt。")

    timestamp = int(time.time())
    date = dt.datetime.utcfromtimestamp(timestamp).strftime("%Y-%m-%d")
    canonical_headers = (
        "content-type:application/json; charset=utf-8\n"
        f"host:{TENCENT_HOST}\n"
        f"x-tc-action:{TENCENT_ACTION.lower()}\n"
    )
    signed_headers = "content-type;host;x-tc-action"
    canonical_request = "\n".join(
        ["POST", "/", "", canonical_headers, signed_headers, sha256_hex(payload)]
    )
    credential_scope = f"{date}/{TENCENT_SERVICE}/tc3_request"
    string_to_sign = "\n".join(
        ["TC3-HMAC-SHA256", str(timestamp), credential_scope, sha256_hex(canonical_request)]
    )

    secret_date = hmac_sha256(("TC3" + secret_key).encode("utf-8"), date)
    secret_service = hmac_sha256(secret_date, TENCENT_SERVICE)
    secret_signing = hmac_sha256(secret_service, "tc3_request")
    signature = hmac.new(secret_signing, string_to_sign.encode("utf-8"), hashlib.sha256).hexdigest()
    authorization = (
        "TC3-HMAC-SHA256 "
        f"Credential={secret_id}/{credential_scope}, "
        f"SignedHeaders={signed_headers}, "
        f"Signature={signature}"
    )

    headers = {
        "Authorization": authorization,
        "Content-Type": "application/json; charset=utf-8",
        "Host": TENCENT_HOST,
        "X-TC-Action": TENCENT_ACTION,
        "X-TC-Timestamp": str(timestamp),
        "X-TC-Version": TENCENT_VERSION,
        "X-TC-Region": tencent_region(),
    }
    token = tencent_token()
    if token:
        headers["X-TC-Token"] = token
    return headers


def wav_duration_seconds(audio_bytes: bytes) -> float:
    with wave.open(io.BytesIO(audio_bytes), "rb") as wav_file:
        frames = wav_file.getnframes()
        frame_rate = wav_file.getframerate()
        if frame_rate <= 0:
            raise ValueError("WAV 采样率无效。")
        return frames / frame_rate


def build_sentence_request_payload(audio_bytes: bytes) -> dict[str, Any]:
    request_payload: dict[str, Any] = {
        "ProjectId": 0,
        "SubServiceType": 2,
        "EngSerViceType": get_env("TENCENT_ASR_ENGINE", default="16k_zh_medical"),
        "SourceType": 1,
        "VoiceFormat": "wav",
        "UsrAudioKey": uuid.uuid4().hex,
        "Data": base64.b64encode(audio_bytes).decode("ascii"),
        "DataLen": len(audio_bytes),
        "FilterDirty": 0,
        "FilterModal": 0,
        "FilterPunc": int(os.getenv("TENCENT_ASR_FILTER_PUNC", "0")),
        "ConvertNumMode": 1,
        "WordInfo": 0,
    }
    hotword_id = os.getenv("TENCENT_ASR_HOTWORD_ID")
    if hotword_id:
        request_payload["HotwordId"] = hotword_id
    return request_payload


def call_tencent_http(request_payload: dict[str, Any]) -> dict[str, Any]:
    payload = json.dumps(request_payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    request = urllib.request.Request(
        TENCENT_ENDPOINT,
        data=payload,
        headers=build_tencent_headers(payload),
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=30) as response:
            response_body = response.read()
    except urllib.error.HTTPError as exc:
        response_body = exc.read()
        raise RuntimeError(parse_tencent_error(response_body) or f"腾讯云接口返回 HTTP {exc.code}") from exc
    except urllib.error.URLError as exc:
        raise RuntimeError(f"无法连接腾讯云接口：{exc.reason}") from exc
    return json.loads(response_body.decode("utf-8"))


def parse_sentence_response(parsed: dict[str, Any]) -> dict[str, Any]:
    response_data = parsed.get("Response", parsed)
    if "Error" in response_data:
        error = response_data["Error"]
        message = error.get("Message") or "腾讯云识别失败。"
        code = error.get("Code")
        raise RuntimeError(f"{code}: {message}" if code else message)
    return response_data


def transcribe_audio(audio_bytes: bytes) -> dict[str, Any]:
    if len(audio_bytes) > MAX_AUDIO_BYTES:
        raise ValueError("音频文件过大，请录制 60 秒以内的短音频。")
    duration = wav_duration_seconds(audio_bytes)
    if duration > MAX_AUDIO_SECONDS:
        raise ValueError("音频超过 60 秒，请重录。")

    request_payload = build_sentence_request_payload(audio_bytes)
    started = time.perf_counter()
    parsed = call_tencent_http(request_payload)
    elapsed_ms = round((time.perf_counter() - started) * 1000)
    response_data = parse_sentence_response(parsed)
    return {
        "text": response_data.get("Result", ""),
        "audioDuration": response_data.get("AudioDuration", duration),
        "latencyMs": elapsed_ms,
        "requestId": response_data.get("RequestId", ""),
        "engine": request_payload["EngSerViceType"],
        "transport": "http",
        "audioBytes": len(audio_bytes),
    }


def parse_tencent_error(response_body: bytes) -> str:
    try:
        parsed = json.loads(response_body.decode("utf-8"))
        error = parsed.get("Response", {}).get("Error", {})
        code = error.get("Code")
        message = error.get("Message")
        if code and message:
            return f"{code}: {message}"
        return message or code or ""
    except Exception:
        return response_body.decode("utf-8", errors="replace")[:500]


def parse_record_text(text: str) -> dict[str, Any]:
    normalized = normalize_text(text)
    fields: dict[str, Any] = {
        "gender": extract_gender(normalized),
        "age": extract_number_field(normalized, "年龄", ("岁", "个月", "月")),
        "weight": extract_number_field(normalized, "体重", ("斤", "公斤", "千克", "kg", "KG")),
        "height": extract_number_field(normalized, "身高", ("厘米", "公分", "cm", "CM")),
        "address": extract_text_between(normalized, "地址", ("联系电话", "电话", "手机", "身高", "金额", "剩余", "还剩")),
        "phone": extract_phone(normalized),
        "remainingSessions": extract_remaining(normalized),
    }
    recharges = extract_recharges(normalized)
    return {"fields": fields, "recharges": recharges}


def normalize_text(text: str) -> str:
    table = str.maketrans(
        "０１２３４５６７８９：，。；（）－",
        "0123456789:,.;()-",
    )
    return re.sub(r"\s+", "", text.translate(table).strip())


def extract_gender(text: str) -> str:
    match = re.search(fr"性别{LABEL_SEP}([男女])", text)
    return match.group(1) if match else ""


def extract_number_field(text: str, label: str, units: tuple[str, ...]) -> str:
    unit_pattern = "|".join(re.escape(unit) for unit in units)
    match = re.search(fr"{label}{LABEL_SEP}({NUMBER_TOKEN})(?:{unit_pattern})?", text, re.IGNORECASE)
    if not match:
        return ""
    value = number_token_to_float(match.group(1))
    if value is None:
        return match.group(1)
    if float(value).is_integer():
        return str(int(value))
    return str(value)


def extract_text_between(text: str, label: str, stop_labels: tuple[str, ...]) -> str:
    stop_pattern = "|".join(re.escape(stop) for stop in stop_labels)
    match = re.search(fr"{label}{LABEL_SEP}(.*?)(?={stop_pattern}|$)", text)
    return match.group(1).strip("，,。.;；") if match else ""


def extract_phone(text: str) -> str:
    match = re.search(fr"(?:联系电话|电话|手机){LABEL_SEP}([0-9零〇一二两三四五六七八九幺\-\s]{{6,22}})", text)
    if not match:
        return ""
    raw = match.group(1)
    digits: list[str] = []
    for char in raw:
        if char.isdigit():
            digits.append(char)
        elif char in CN_DIGITS:
            digits.append(str(CN_DIGITS[char]))
    return "".join(digits)


def extract_remaining(text: str) -> int | None:
    match = re.search(
        fr"(?:目前)?(?:剩余次数|剩余|还剩|剩){LABEL_SEP}(?:为|是|还有|有)?{LABEL_SEP}({NUMBER_TOKEN}){LABEL_SEP}次",
        text,
    )
    if not match:
        return None
    value = number_token_to_int(match.group(1))
    return value


def extract_recharges(text: str) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    scan_recharge_pairs(text, items)

    for match in re.finditer(fr"(?:金额|充值|充){LABEL_SEP}(.*?)(?=金额|充值|充|剩余|还剩|$)", text):
        chunk = match.group(1).strip("，,。.;；")
        if not chunk:
            continue
        if scan_recharge_pairs(chunk, items):
            continue
        amount = extract_amount(chunk)
        date_text = extract_date(chunk)
        add_recharge_item(items, date_text, amount, chunk)
    return items


def scan_recharge_pairs(text: str, items: list[dict[str, Any]]) -> int:
    date_pattern = re.compile(DATE_FRAGMENT)
    amount_pattern = re.compile(fr"({NUM})(?:元|块|块钱)")
    date_matches = list(date_pattern.finditer(text))
    count = 0
    for index, date_match in enumerate(date_matches):
        next_start = date_matches[index + 1].start() if index + 1 < len(date_matches) else len(text)
        segment = text[date_match.start() : next_start]
        amount_match = amount_pattern.search(segment)
        if not amount_match:
            continue
        chunk = segment[: amount_match.end()].strip("，,。.;；")
        date_text = extract_date(date_match.group(0))
        amount = number_token_to_int(amount_match.group(1))
        add_recharge_item(items, date_text, amount, chunk)
        count += 1
    return count


def add_recharge_item(
    items: list[dict[str, Any]],
    date_text: str,
    amount: int | None,
    raw_text: str,
) -> None:
    if amount is None and not date_text:
        return
    if any(item.get("rawText") == raw_text for item in items):
        return
    if any(item.get("date") == date_text and item.get("amount") == amount for item in items):
        return
    items.append(
        {
            "date": date_text,
            "amount": amount,
            "sessions": infer_sessions(amount),
            "rawText": raw_text,
        }
    )


def extract_amount(text: str) -> int | None:
    match = re.search(fr"({NUMBER_TOKEN})(?:元|块|块钱)", text)
    if not match:
        all_numbers = re.findall(NUMBER_TOKEN, text)
        if not all_numbers:
            return None
        match_text = all_numbers[-1]
    else:
        match_text = match.group(1)
    return number_token_to_int(match_text)


def extract_date(text: str) -> str:
    year_match = re.search(fr"({NUMBER_TOKEN})年({NUMBER_TOKEN})月({NUMBER_TOKEN})(?:日|号)?", text)
    if year_match:
        return format_date_parts(year_match.group(1), year_match.group(2), year_match.group(3))

    compact_year_match = re.search(r"([0-9]{4})([0-9]{1,2})月([0-9]{1,2})(?:日|号)?", text)
    if compact_year_match:
        return format_date_parts(
            compact_year_match.group(1),
            compact_year_match.group(2),
            compact_year_match.group(3),
        )

    month_match = re.search(fr"({NUMBER_TOKEN})月({NUMBER_TOKEN})(?:日|号)?", text)
    if month_match:
        month = number_token_to_int(month_match.group(1))
        day = number_token_to_int(month_match.group(2))
        if month is not None and day is not None and is_valid_month_day(month, day):
            return f"{month:02d}-{day:02d}"
    return ""


def format_date_parts(year_text: str, month_text: str, day_text: str) -> str:
    year = number_token_to_int(year_text)
    month = number_token_to_int(month_text)
    day = number_token_to_int(day_text)
    if year is None or month is None or day is None:
        return ""
    if year < 100:
        year += 2000
    if not is_valid_month_day(month, day):
        return ""
    return f"{year:04d}-{month:02d}-{day:02d}"


def is_valid_month_day(month: int, day: int) -> bool:
    return 1 <= month <= 12 and 1 <= day <= 31


def infer_sessions(amount: int | None) -> int | None:
    if amount is None:
        return None
    if amount % 90 == 0:
        return amount // 90
    if amount == 900:
        return 10
    return None


def number_token_to_float(token: str) -> float | None:
    token = token.strip()
    if not token:
        return None
    if re.fullmatch(r"[0-9]+(?:\.[0-9]+)?", token):
        return float(token)
    if "点" in token:
        integer, fraction = token.split("点", 1)
        int_value = number_token_to_int(integer) if integer else 0
        if int_value is None:
            return None
        fraction_digits = "".join(str(CN_DIGITS.get(char, "")) for char in fraction)
        if not fraction_digits:
            return float(int_value)
        return float(f"{int_value}.{fraction_digits}")
    int_value = number_token_to_int(token)
    return float(int_value) if int_value is not None else None


def number_token_to_int(token: str) -> int | None:
    token = token.strip()
    if not token:
        return None
    if re.fullmatch(r"[0-9]+(?:\.[0-9]+)?", token):
        return int(float(token))
    if all(char in CN_DIGITS for char in token):
        return int("".join(str(CN_DIGITS[char]) for char in token))

    total = 0
    section = 0
    number = 0
    for char in token:
        if char in CN_DIGITS:
            number = CN_DIGITS[char]
            continue
        unit = CN_UNITS.get(char)
        if not unit:
            continue
        if unit == 10000:
            section = (section + number) * unit
            total += section
            section = 0
        else:
            section += (number or 1) * unit
        number = 0
    return total + section + number


SIGNATURE_KIND_TO_FIELD = {
    "directory": "directorySignature",
    "case": "caseSignature",
    "visit": "visitSignature",
    "flow": "flowSignature",
}


def load_signature_manifest() -> dict[str, Any]:
    manifest = json.loads(SIGNATURE_MANIFEST_PATH.read_text(encoding="utf-8"))
    if not isinstance(manifest, dict):
        raise ValueError("signature manifest is invalid")
    return apply_signature_bindings(manifest)


def load_signature_bindings() -> dict[str, Any]:
    if not SIGNATURE_BINDINGS_PATH.exists():
        return {"updatedAt": "", "patients": {}}
    parsed = json.loads(SIGNATURE_BINDINGS_PATH.read_text(encoding="utf-8"))
    if not isinstance(parsed, dict):
        return {"updatedAt": "", "patients": {}}
    patients = parsed.get("patients")
    if not isinstance(patients, dict):
        parsed["patients"] = {}
    return parsed


def save_signature_bindings(bindings: dict[str, Any]) -> None:
    SIGNATURE_PATIENT_DIR.mkdir(parents=True, exist_ok=True)
    temp_path = SIGNATURE_BINDINGS_PATH.with_suffix(".tmp")
    temp_path.write_text(
        json.dumps(bindings, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    temp_path.replace(SIGNATURE_BINDINGS_PATH)


def apply_signature_bindings(manifest: dict[str, Any]) -> dict[str, Any]:
    try:
        bindings = load_signature_bindings()
    except Exception:
        bindings = {"updatedAt": "", "patients": {}}
    bound_patients = bindings.get("patients") or {}
    if not isinstance(bound_patients, dict):
        bound_patients = {}

    for item in manifest.get("items") or []:
        if not isinstance(item, dict):
            continue
        patient_binding = bound_patients.get(str(item.get("patientId")))
        if not isinstance(patient_binding, dict):
            item["hasDirectorySignature"] = bool(item.get("directorySignature"))
            continue
        for field in SIGNATURE_KIND_TO_FIELD.values():
            url = patient_binding.get(field)
            if isinstance(url, str) and url:
                item[field] = url
        item["hasDirectorySignature"] = bool(item.get("directorySignature"))
        if patient_binding.get("updatedAt"):
            item["electronicSignatureUpdatedAt"] = patient_binding["updatedAt"]

    if bindings.get("updatedAt"):
        manifest["bindingUpdatedAt"] = bindings["updatedAt"]
    return manifest


def signature_public_url(path: Path) -> str:
    root = SIGNATURE_PATIENT_DIR.resolve()
    resolved = path.resolve()
    if resolved != root and root not in resolved.parents:
        raise ValueError("signature path escaped storage directory")
    relative = resolved.relative_to(root)
    return "/doc_patient/signature_patient/" + relative.as_posix()


def decode_signature_png(payload: dict[str, Any]) -> bytes:
    image_data = str(payload.get("imageData") or "")
    if "," in image_data:
        header, image_base64 = image_data.split(",", 1)
        if "image/png" not in header:
            raise ValueError("only png signatures are supported")
    else:
        image_base64 = image_data
    if not image_base64:
        raise ValueError("missing signature image")

    image_bytes = base64.b64decode(image_base64, validate=True)
    if len(image_bytes) > MAX_SIGNATURE_IMAGE_BYTES:
        raise ValueError("signature image is too large")
    if not image_bytes.startswith(b"\x89PNG\r\n\x1a\n"):
        raise ValueError("signature image must be a png")
    return image_bytes


def save_electronic_signature(payload: dict[str, Any]) -> dict[str, Any]:
    patient_id = int(payload.get("patientId") or 0)
    patient = get_patient(patient_id)
    if not patient:
        raise ValueError("patient not found")

    kind = str(payload.get("kind") or "visit").strip()
    field = SIGNATURE_KIND_TO_FIELD.get(kind)
    if not field:
        raise ValueError("signature kind must be directory, case, visit, or flow")
    adjustment_id = optional_int(payload.get("adjustmentId"))

    image_bytes = decode_signature_png(payload)
    saved_at = now_text()
    timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    order = int(patient.get("order") or patient_id)
    target_dir = ELECTRONIC_SIGNATURE_DIR / kind
    target_dir.mkdir(parents=True, exist_ok=True)
    target_path = target_dir / f"patient_{order:03d}_{kind}_{timestamp}.png"
    target_path.write_bytes(image_bytes)
    url = signature_public_url(target_path)

    bindings = load_signature_bindings()
    bindings["updatedAt"] = saved_at
    bound_patients = bindings.setdefault("patients", {})
    patient_binding = bound_patients.setdefault(str(patient_id), {})
    patient_binding[field] = url
    patient_binding["updatedAt"] = saved_at

    history = patient_binding.setdefault("history", [])
    if not isinstance(history, list):
        history = []
        patient_binding["history"] = history
    history.append(
        {
            "kind": kind,
            "field": field,
            "url": url,
            "adjustmentId": adjustment_id or "",
            "signerName": str(payload.get("signerName") or "").strip(),
            "note": str(payload.get("note") or "").strip(),
            "savedAt": saved_at,
        }
    )
    del history[:-50]
    save_signature_bindings(bindings)

    if adjustment_id:
        mark_adjustment_signed(
            patient_id=patient_id,
            adjustment_id=adjustment_id,
            url=url,
            saved_at=saved_at,
            signer_name=str(payload.get("signerName") or "").strip(),
            note=str(payload.get("note") or "").strip(),
        )

    return {
        "patientId": patient_id,
        "kind": kind,
        "field": field,
        "url": url,
        "filePath": str(target_path),
        "savedAt": saved_at,
        "adjustmentId": adjustment_id or "",
    }


def mark_adjustment_signed(
    patient_id: int,
    adjustment_id: int,
    url: str,
    saved_at: str,
    signer_name: str,
    note: str,
) -> None:
    with connect_db() as conn:
        row = conn.execute(
            """
            SELECT id
            FROM session_adjustments
            WHERE id = ? AND patient_id = ?
            """,
            (adjustment_id, patient_id),
        ).fetchone()
        if not row:
            raise ValueError("没有找到要绑定签名的流水")
        conn.execute(
            """
            UPDATE session_adjustments
            SET signature_status = 'signed',
                signature_url = ?,
                signature_saved_at = ?,
                signature_signer = ?,
                signature_note = ?
            WHERE id = ? AND patient_id = ?
            """,
            (url, saved_at, signer_name, note, adjustment_id, patient_id),
        )
        conn.commit()


def list_pending_signature_adjustments(query: dict[str, list[str]]) -> dict[str, Any]:
    range_type = first_query_value(query, "range") or "month"
    if range_type not in {"day", "month", "year"}:
        raise ValueError("筛选范围只能是 day、month 或 year")
    date_value = first_query_value(query, "date")
    if not date_value:
        now = dt.datetime.now()
        date_value = now.strftime("%Y-%m") if range_type == "month" else now.strftime("%Y-%m-%d")
    date_prefix = normalize_summary_date(range_type, date_value)
    prefix_length = len(date_prefix)
    patient_id = optional_int(first_query_value(query, "patientId"))
    therapist = first_query_value(query, "therapist")
    if therapist and therapist not in THERAPISTS:
        raise ValueError("师傅筛选无效")

    clauses = [
        f"substr(a.occurred_at, 1, {prefix_length}) = ?",
        "COALESCE(a.signature_status, 'pending') = 'pending'",
        "a.voided_at IS NULL",
        "a.correction_of_adjustment_id IS NULL",
    ]
    params: list[Any] = [date_prefix]
    if patient_id:
        clauses.append("a.patient_id = ?")
        params.append(patient_id)
    if therapist:
        clauses.append("a.operation = 'decrease'")
        clauses.append("a.therapist = ?")
        params.append(therapist)

    with connect_db() as conn:
        rows = conn.execute(
            f"""
            SELECT a.*, p.name AS patient_name, p.import_order AS patient_order,
                   p.address AS patient_address, p.record_no AS patient_record_no
            FROM session_adjustments a
            JOIN patients p ON p.id = a.patient_id
            WHERE {" AND ".join(clauses)}
            ORDER BY a.occurred_at DESC, a.id DESC
            """,
            params,
        ).fetchall()
    records = [row_to_summary_adjustment(row) for row in rows]
    return {
        "filters": {
            "range": range_type,
            "date": date_prefix,
            "patientId": patient_id or "",
            "therapist": therapist or "",
        },
        "records": records,
        "summary": {
            "recordCount": len(records),
            "patientCount": len({item["patientId"] for item in records}),
            "rechargeAmount": sum_number(item["amount"] for item in records if item["operation"] == "increase"),
            "rechargeSessions": sum_number(item["sessions"] for item in records if item["operation"] == "increase"),
            "massageSessions": sum_number(item["sessions"] for item in records if item["operation"] == "decrease"),
        },
    }


def save_bulk_flow_signature(payload: dict[str, Any]) -> dict[str, Any]:
    raw_ids = payload.get("adjustmentIds")
    if not isinstance(raw_ids, list):
        raise ValueError("请选择要补签的流水")
    adjustment_ids = sorted({int(value) for value in raw_ids if int(value) > 0})
    if not adjustment_ids:
        raise ValueError("请选择要补签的流水")
    if len(adjustment_ids) > 100:
        raise ValueError("一次批量补签最多选择 100 条流水")

    image_bytes = decode_signature_png(payload)
    signer_name = str(payload.get("signerName") or "").strip()
    note = str(payload.get("note") or "").strip()
    saved_at = now_text()
    timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    target_dir = ELECTRONIC_SIGNATURE_DIR / "flow_batch"
    target_dir.mkdir(parents=True, exist_ok=True)
    target_path = target_dir / f"bulk_flow_{timestamp}.png"
    target_path.write_bytes(image_bytes)
    url = signature_public_url(target_path)

    placeholders = ",".join("?" for _ in adjustment_ids)
    with connect_db() as conn:
        rows = conn.execute(
            f"""
            SELECT a.*, p.name AS patient_name, p.import_order AS patient_order,
                   p.address AS patient_address, p.record_no AS patient_record_no
            FROM session_adjustments a
            JOIN patients p ON p.id = a.patient_id
            WHERE a.id IN ({placeholders})
            ORDER BY a.occurred_at DESC, a.id DESC
            """,
            adjustment_ids,
        ).fetchall()
        if len(rows) != len(adjustment_ids):
            raise ValueError("部分流水不存在")
        for row in rows:
            if row["voided_at"] or row["correction_of_adjustment_id"]:
                raise ValueError("已冲正或冲正流水不能批量补签")
            if (row["signature_status"] or "pending") != "pending":
                raise ValueError("只能批量补签待签名流水")

        cursor = conn.execute(
            """
            INSERT INTO signature_batches (signature_url, signer, note, adjustment_count, signed_at, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (url, signer_name, note, len(rows), saved_at, saved_at),
        )
        batch_id = int(cursor.lastrowid)
        conn.execute(
            f"""
            UPDATE session_adjustments
            SET signature_status = 'signed',
                signature_url = ?,
                signature_saved_at = ?,
                signature_signer = ?,
                signature_note = ?,
                signature_batch_id = ?
            WHERE id IN ({placeholders})
            """,
            [url, saved_at, signer_name, note, batch_id, *adjustment_ids],
        )
        append_bulk_signature_bindings(rows, url, saved_at, signer_name, note, batch_id)
        conn.commit()

    return {
        "batch": {
            "id": batch_id,
            "url": url,
            "signerName": signer_name,
            "note": note,
            "savedAt": saved_at,
            "adjustmentCount": len(adjustment_ids),
        },
        "records": [get_session_adjustment(adjustment_id) for adjustment_id in adjustment_ids],
    }


def append_bulk_signature_bindings(
    rows: list[sqlite3.Row],
    url: str,
    saved_at: str,
    signer_name: str,
    note: str,
    batch_id: int,
) -> None:
    bindings = load_signature_bindings()
    bindings["updatedAt"] = saved_at
    bound_patients = bindings.setdefault("patients", {})
    for row in rows:
        patient_id = str(row["patient_id"])
        patient_binding = bound_patients.setdefault(patient_id, {})
        field = "visitSignature" if row["operation"] == "decrease" else "flowSignature"
        kind = "visit" if row["operation"] == "decrease" else "flow"
        patient_binding[field] = url
        patient_binding["updatedAt"] = saved_at
        history = patient_binding.setdefault("history", [])
        if not isinstance(history, list):
            history = []
            patient_binding["history"] = history
        history.append(
            {
                "kind": kind,
                "field": field,
                "url": url,
                "adjustmentId": row["id"],
                "bulkSignatureBatchId": batch_id,
                "signerName": signer_name,
                "note": note,
                "savedAt": saved_at,
            }
        )
        del history[:-50]
    save_signature_bindings(bindings)


def save_test_signature(payload: dict[str, Any]) -> dict[str, Any]:
    kind = str(payload.get("kind") or "visit").strip()
    if kind not in SIGNATURE_KIND_TO_FIELD:
        raise ValueError("signature kind must be directory, case, or visit")

    image_bytes = decode_signature_png(payload)
    saved_at = now_text()
    timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    target_dir = TEST_SIGNATURE_DIR / kind
    target_dir.mkdir(parents=True, exist_ok=True)
    target_path = target_dir / f"test_{kind}_{timestamp}.png"
    target_path.write_bytes(image_bytes)
    url = signature_public_url(target_path)
    return {
        "kind": kind,
        "url": url,
        "filePath": str(target_path),
        "savedAt": saved_at,
        "signerName": str(payload.get("signerName") or "").strip(),
        "note": str(payload.get("note") or "").strip(),
    }


def list_test_signatures(limit: int = 20) -> list[dict[str, Any]]:
    if not TEST_SIGNATURE_DIR.exists():
        return []
    files = sorted(
        TEST_SIGNATURE_DIR.glob("*/*.png"),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    items: list[dict[str, Any]] = []
    for path in files[:limit]:
        modified_at = dt.datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S")
        items.append(
            {
                "kind": path.parent.name,
                "url": signature_public_url(path),
                "filePath": str(path),
                "savedAt": modified_at,
            }
        )
    return items


class TuinaHandler(SimpleHTTPRequestHandler):
    def translate_path(self, path: str) -> str:
        clean_path = urllib.parse.urlparse(path).path
        if clean_path in ("/", "/index.html"):
            return str(STATIC_DIR / "index.html")
        if clean_path in ("/patient-search", "/patient-search.html"):
            return str(STATIC_DIR / "patient-search.html")
        if clean_path in ("/patient-sessions", "/patient-sessions.html"):
            return str(STATIC_DIR / "patient-sessions.html")
        if clean_path in ("/signature-browser", "/signature-browser.html"):
            return str(STATIC_DIR / "signature-browser.html")
        if clean_path in ("/signature-review", "/signature-review.html"):
            return str(STATIC_DIR / "signature-review.html")
        if clean_path in ("/signature-pad", "/signature-pad.html"):
            return str(STATIC_DIR / "signature-pad.html")
        if clean_path in ("/signature-test", "/signature-test.html"):
            return str(STATIC_DIR / "signature-test.html")
        if clean_path.startswith("/doc_patient/signature_patient/"):
            relative = urllib.parse.unquote(clean_path.removeprefix("/doc_patient/signature_patient/"))
            candidate = (SIGNATURE_PATIENT_DIR / relative).resolve()
            root = SIGNATURE_PATIENT_DIR.resolve()
            if candidate == root or root in candidate.parents:
                return str(candidate)
            return str(STATIC_DIR / "__not_found__")
        return str(STATIC_DIR / clean_path.lstrip("/"))

    def end_headers(self) -> None:
        self.send_header("X-Content-Type-Options", "nosniff")
        self.send_header("Cache-Control", "no-store")
        super().end_headers()

    def do_GET(self) -> None:
        parsed_path = urllib.parse.urlparse(self.path)
        if parsed_path.path == "/api/config":
            json_response(self, HTTPStatus.OK, {"ok": True, "config": public_config()})
            return
        if parsed_path.path == "/api/signature-manifest":
            if not SIGNATURE_MANIFEST_PATH.exists():
                json_response(
                    self,
                    HTTPStatus.NOT_FOUND,
                    {"ok": False, "error": "signature manifest not generated"},
                )
                return
            try:
                manifest = load_signature_manifest()
                json_response(self, HTTPStatus.OK, {"ok": True, "manifest": manifest})
            except Exception as exc:
                json_response(self, HTTPStatus.BAD_REQUEST, {"ok": False, "error": str(exc)})
            return
        if parsed_path.path == "/api/test-signatures":
            try:
                json_response(
                    self,
                    HTTPStatus.OK,
                    {"ok": True, "signatures": list_test_signatures()},
                )
            except Exception as exc:
                json_response(self, HTTPStatus.BAD_REQUEST, {"ok": False, "error": str(exc)})
            return
        if parsed_path.path == "/api/patients":
            init_database()
            json_response(self, HTTPStatus.OK, {"ok": True, "patients": list_patients()})
            return
        if parsed_path.path == "/api/session-summary":
            init_database()
            try:
                query = urllib.parse.parse_qs(parsed_path.query, keep_blank_values=True)
                json_response(self, HTTPStatus.OK, {"ok": True, **build_session_summary(query)})
            except Exception as exc:
                json_response(self, HTTPStatus.BAD_REQUEST, {"ok": False, "error": str(exc)})
            return
        if parsed_path.path == "/api/settlements":
            init_database()
            try:
                json_response(self, HTTPStatus.OK, {"ok": True, **list_settlements()})
            except Exception as exc:
                json_response(self, HTTPStatus.BAD_REQUEST, {"ok": False, "error": str(exc)})
            return
        if parsed_path.path.startswith("/api/settlements/"):
            init_database()
            try:
                settlement_id = int(parsed_path.path.rsplit("/", 1)[1])
                settlement = get_settlement(settlement_id)
                if not settlement:
                    json_response(self, HTTPStatus.NOT_FOUND, {"ok": False, "error": "没有找到这份月结单"})
                else:
                    json_response(self, HTTPStatus.OK, {"ok": True, "settlement": settlement})
            except ValueError as exc:
                json_response(self, HTTPStatus.BAD_REQUEST, {"ok": False, "error": str(exc)})
            return
        if parsed_path.path == "/api/pending-signatures":
            init_database()
            try:
                query = urllib.parse.parse_qs(parsed_path.query, keep_blank_values=True)
                json_response(self, HTTPStatus.OK, {"ok": True, **list_pending_signature_adjustments(query)})
            except Exception as exc:
                json_response(self, HTTPStatus.BAD_REQUEST, {"ok": False, "error": str(exc)})
            return
        if parsed_path.path == "/api/backup-status":
            json_response(self, HTTPStatus.OK, {"ok": True, "backup": get_backup_status()})
            return
        if parsed_path.path.startswith("/api/session-page/"):
            init_database()
            try:
                patient_id = int(parsed_path.path.rsplit("/", 1)[1])
                json_response(self, HTTPStatus.OK, {"ok": True, **get_session_page(patient_id)})
            except Exception as exc:
                json_response(self, HTTPStatus.BAD_REQUEST, {"ok": False, "error": str(exc)})
            return
        if parsed_path.path.startswith("/api/patients/"):
            init_database()
            try:
                patient_id = int(parsed_path.path.rsplit("/", 1)[1])
            except ValueError:
                json_response(self, HTTPStatus.BAD_REQUEST, {"ok": False, "error": "人名 ID 无效。"})
                return
            patient = get_patient(patient_id)
            if not patient:
                json_response(self, HTTPStatus.NOT_FOUND, {"ok": False, "error": "没有找到这个人名。"})
                return
            json_response(self, HTTPStatus.OK, {"ok": True, "patient": patient})
            return
        if parsed_path.path == "/api/export.csv":
            init_database()
            body = export_csv()
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "text/csv; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.send_header(
                "Content-Disposition",
                'attachment; filename="tuina_records.csv"',
            )
            self.end_headers()
            self.wfile.write(body)
            return
        return super().do_GET()

    def do_POST(self) -> None:
        parsed_path = urllib.parse.urlparse(self.path)
        if parsed_path.path == "/api/settlements":
            try:
                init_database()
                request_json = read_json_body(self)
                settlement = create_settlement(request_json)
                json_response(self, HTTPStatus.OK, {"ok": True, "settlement": settlement})
            except Exception as exc:
                json_response(self, HTTPStatus.BAD_REQUEST, {"ok": False, "error": str(exc)})
            return

        if parsed_path.path.startswith("/api/settlements/") and parsed_path.path.endswith("/revoke"):
            try:
                init_database()
                settlement_id = int(parsed_path.path.strip("/").split("/")[-2])
                request_json = read_json_body(self)
                result = revoke_settlement(settlement_id, request_json)
                json_response(self, HTTPStatus.OK, {"ok": True, **result})
            except Exception as exc:
                json_response(self, HTTPStatus.BAD_REQUEST, {"ok": False, "error": str(exc)})
            return

        if parsed_path.path == "/api/parse":
            try:
                request_json = read_json_body(self)
                text = str(request_json.get("text") or "")
                json_response(self, HTTPStatus.OK, {"ok": True, "parsed": parse_record_text(text)})
            except Exception as exc:
                json_response(self, HTTPStatus.BAD_REQUEST, {"ok": False, "error": str(exc)})
            return

        if parsed_path.path == "/api/transcribe":
            try:
                request_json = read_json_body(self)
                audio_base64 = str(request_json.get("audioBase64") or "")
                if "," in audio_base64:
                    audio_base64 = audio_base64.split(",", 1)[1]
                if not audio_base64:
                    raise ValueError("没有收到录音数据。")
                audio_bytes = base64.b64decode(audio_base64, validate=True)
                result = transcribe_audio(audio_bytes)
                json_response(self, HTTPStatus.OK, {"ok": True, **result})
            except Exception as exc:
                json_response(self, HTTPStatus.BAD_REQUEST, {"ok": False, "error": str(exc)})
            return

        if parsed_path.path == "/api/patients":
            try:
                init_database()
                request_json = read_json_body(self)
                patient = create_patient(request_json)
                json_response(self, HTTPStatus.OK, {"ok": True, "patient": patient})
            except Exception as exc:
                json_response(self, HTTPStatus.BAD_REQUEST, {"ok": False, "error": str(exc)})
            return

        if parsed_path.path == "/api/signatures":
            try:
                init_database()
                request_json = read_json_body(self)
                signature = save_electronic_signature(request_json)
                json_response(self, HTTPStatus.OK, {"ok": True, "signature": signature})
            except Exception as exc:
                json_response(self, HTTPStatus.BAD_REQUEST, {"ok": False, "error": str(exc)})
            return

        if parsed_path.path == "/api/bulk-signatures":
            try:
                init_database()
                request_json = read_json_body(self)
                result = save_bulk_flow_signature(request_json)
                json_response(self, HTTPStatus.OK, {"ok": True, **result})
            except Exception as exc:
                json_response(self, HTTPStatus.BAD_REQUEST, {"ok": False, "error": str(exc)})
            return

        if parsed_path.path == "/api/backup-now":
            try:
                set_backup_status("running", "正在手动增量备份", configured=True)
                result = perform_cloud_backup(force=True)
                set_backup_status(
                    "success",
                    "手动增量备份已保存",
                    configured=True,
                    lastSuccessAt=result.get("lastSuccessAt", ""),
                    lastRemoteName=result.get("lastRemoteName", ""),
                    lastIncludedSignatures=result.get("lastIncludedSignatures", 0),
                    lastPackageBytes=result.get("lastPackageBytes", 0),
                )
                json_response(self, HTTPStatus.OK, {"ok": True, "backup": get_backup_status()})
            except Exception as exc:
                set_backup_status("error", backup_error_message(exc), configured=True)
                json_response(self, HTTPStatus.BAD_REQUEST, {"ok": False, "error": str(exc), "backup": get_backup_status()})
            return

        if parsed_path.path == "/api/test-signatures":
            try:
                request_json = read_json_body(self)
                signature = save_test_signature(request_json)
                json_response(self, HTTPStatus.OK, {"ok": True, "signature": signature})
            except Exception as exc:
                json_response(self, HTTPStatus.BAD_REQUEST, {"ok": False, "error": str(exc)})
            return

        if parsed_path.path.startswith("/api/session-adjustments/"):
            try:
                init_database()
                request_json = read_json_body(self)
                if parsed_path.path.endswith("/reverse"):
                    adjustment_id = int(parsed_path.path.strip("/").split("/")[-2])
                    result = reverse_session_adjustment(adjustment_id, request_json)
                else:
                    patient_id = int(parsed_path.path.rsplit("/", 1)[1])
                    result = apply_session_adjustment(patient_id, request_json)
                json_response(self, HTTPStatus.OK, {"ok": True, **result})
            except Exception as exc:
                json_response(self, HTTPStatus.BAD_REQUEST, {"ok": False, "error": str(exc)})
            return

        if parsed_path.path.startswith("/api/patients/"):
            try:
                init_database()
                patient_id = int(parsed_path.path.rsplit("/", 1)[1])
                request_json = read_json_body(self)
                saved = save_patient(patient_id, request_json)
                json_response(self, HTTPStatus.OK, {"ok": True, "patient": saved})
            except Exception as exc:
                json_response(self, HTTPStatus.BAD_REQUEST, {"ok": False, "error": str(exc)})
            return

        json_response(self, HTTPStatus.NOT_FOUND, {"ok": False, "error": "接口不存在。"})

    def guess_type(self, path: str) -> str:
        if path.endswith(".js"):
            return "text/javascript"
        if path.endswith(".css"):
            return "text/css"
        return mimetypes.guess_type(path)[0] or "application/octet-stream"

    def log_message(self, format: str, *args: Any) -> None:
        sys.stderr.write(f"[tuina] {self.address_string()} {format % args}\n")


class ThreadedTCPServer(socketserver.ThreadingMixIn, socketserver.TCPServer):
    allow_reuse_address = True
    daemon_threads = True


def run_server(host: str, port: int) -> None:
    load_dotenv()
    load_tencent_key_file()
    load_backup_key_file()
    init_database()
    start_cloud_backup_worker()
    with ThreadedTCPServer((host, port), TuinaHandler) as server:
        print(f"Tuina input system: http://{host}:{port}", flush=True)
        print(f"Database: {DB_PATH}", flush=True)
        print("Press Ctrl+C to stop.", flush=True)
        server.serve_forever()


def main() -> None:
    parser = argparse.ArgumentParser(description="Local Tuina patient input system.")
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", type=int, default=8776)
    args = parser.parse_args()
    run_server(args.host, args.port)


if __name__ == "__main__":
    main()
